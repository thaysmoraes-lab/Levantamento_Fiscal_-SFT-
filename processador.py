#!/usr/bin/env python3
"""
CONFIGURADOR DE TRIBUTOS PROTHEUS — Lógica de Processamento SFT
================================================================
Autor: Claude (sessão Thays Caroline - 16/05/2026)
Versão: v22

Gera planilha de apoio com Perfis e Regras para o Configurador de Tributos
a partir da tabela SFT (Livros Fiscais) do TOTVS Protheus.

Impostos mapeados:
  - ICMS + PIS/COFINS + ISS
  - Retenções (PR, CR, SR, IR, NR, TR)
  - PROTEGE GO (PG)
  - ICMS Complementar DIFAL (CD) e Base Destino (CB)
  - IPI (PI)

NOVO NA v22:
  - PROMPT INTERATIVO no início para capturar perfil da empresa.
    Perguntas estratégicas (segmento, regime tributário, operações comuns)
    eliminam CBENEFs incompatíveis (ex.: indústria aeroespacial, REPORTO,
    transporte trilhos, doações, etc).
  - Coluna 'APLICÁVEL?' (SIM / TALVEZ / NÃO) — todos os CBENEFs ficam
    listados (transparência), mas marcados conforme aplicabilidade.
  - Ordem: SIM primeiro, depois TALVEZ, depois NÃO no final.
  - Resposta padrão (Enter) é a mais conservadora (mantém o código).

NOVO NA v12:
  - 6 retenções via função genérica build_retencao_tables():
      PR = PIS Retido       (Base Pis / Aliq. PIS.1 / Valor Pis)
      CR = COFINS Retido    (Base Cofins / Aliq. Cofins / Valor Cofins)
      SR = CSLL Retido      (Base CSLL.1 / Aliq. CSLL.1 / Valor CSLL.1)
      IR = IRRF             (Base IRRF / Aliq. IRRF / Valor IRRF)
      NR = INSS             (Base INSS / Aliq. INSS / Valor INSS)
      TR = ISS Retido       (flag ISS Retido=1 + Base/Alíq/Valor ICMS)
  - Filtro de retenção: Base > 0 (ISS Retido usa flag == 1)
  - Cada retenção = perfil independente, 4 perfis + 2 regras
  - Tema visual: roxo #5B2C6F

Dependências: pandas, openpyxl, xlrd
"""

import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Classificação manual dos 276 CBENEFs de GO (categorias + capítulos NCM + flags)
# Arquivo separado para manutenibilidade
try:
    from cbenef_classificacao_GO import CBENEF_GO_CLASSIFICACAO, PERFIL_PERGUNTAS
except ImportError:
    CBENEF_GO_CLASSIFICACAO = {}
    PERFIL_PERGUNTAS = []
    print("⚠ cbenef_classificacao_GO.py não encontrado — filtragem por categoria desativada")

# Mapeamento estado → dicionário de classificação
CBENEF_CLASSIFICACAO_POR_ESTADO = {
    'GO': CBENEF_GO_CLASSIFICACAO,
    # 'SP': CBENEF_SP_CLASSIFICACAO,  # quando adicionar outro estado
}

# ============================================================
# CONFIGURAÇÃO — paths relativos a este arquivo
# ============================================================
import os
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.path.join(_BASE_DIR, 'data')

SFT_FILE     = None   # ★ definido em tempo de execução pelo app
SFT_SHEET    = None   # None = detectar automaticamente
CFOP_FILE    = os.path.join(_DATA_DIR, 'Tabela_CFOP.xlsx')
MODELO_FILE  = os.path.join(_DATA_DIR, 'Modelo_planilha_apoio.xlsx')
OUTPUT_FILE  = 'Planilha_Apoio_Configurador_Tributos.xlsx'

# ============================================================
# MAPEAMENTO FILIAL → ESTADO  ★ EDITAR CONFORME A SFT ★
# ============================================================
# Adicione UMA ENTRADA PARA CADA FILIAL que aparece na sua SFT.
# A chave pode ser:
#   - número:  1, 2, 10101
#   - string:  '1', '01', '0101', '10101'
#   - vazio:   '' (quando a coluna Filial está em branco)
# O script normaliza tudo. Recomendação: usar string para evitar problemas.
#
# Se uma filial da SFT não estiver mapeada aqui, o script avisa e usa 'XX'.
FILIAL_STATE = {
    # === Exemplos — REMOVA ou EDITE conforme sua realidade ===
    '10101': 'GO',     # SFT original (filial única)
    '1':     'GO',     # SFT nova com filial '1'
    '2':     'GO',     # SFT nova com filial '2'
    '01':    'GO',     # variante zero-padded
    '0101':  'GO',     # variante 4 dígitos
    '':      'GO',     # filial em branco → default
}

# ============================================================
# ALIAS DE COLUNAS (compatibilidade entre versões de SFT)
# ============================================================
# Mapeia "nome esperado pelo código" → lista de possíveis nomes na SFT.
# O script vai usar o primeiro nome da lista que existir no DataFrame.
# Se nenhum existir, a coluna é tratada como ausente (valor zero/vazio).
COL_ALIASES = {
    'CST Pis':       ['CST Pis', 'Sit.Trib.PIS', 'CST PIS', 'CSTPIS'],
    'CST COF':       ['CST COF', 'Sit.Trib.COF', 'CST COFINS', 'CSTCOF'],
    'Cód. Serviço':  ['Cód. Serviço', 'Cod. Servico', 'Cod Serviço'],
    'C. Trib. Mun':  ['C. Trib. Mun', 'C.Trib.Mun', 'Cod Trib Mun'],
    'Sit.Tribut.':   ['Sit.Tribut.', 'Sit Tribut', 'CST ICMS', 'CST'],
}

# CFOPs de prestação/aquisição de serviço (ISS) — desviados do ICMS
CFOP_ISS = {'1933', '2933', '5933', '6933'}

# ============================================================
# CBENEF — Tabelas de Código de Benefício Fiscal por estado
# ============================================================
# Cada estado pode exigir CBENEF próprio quando há isenção/redução de ICMS.
# Para cada estado em que sua empresa opera, adicione uma entrada apontando
# para o arquivo Excel respectivo (formato esperado: aba 'CBENEF x CFOP'
# com colunas 'CÓDIGO CBENEF', 'CST', 'TIPO BENEFÍCIO', 'CFOP(s) INDICADO(s)').
#
# Se um estado não estiver nesta tabela, a sugestão fica em branco para
# cenários daquela filial — sem erro.
CBENEF_TABLES = {
    'GO': os.path.join(_DATA_DIR, 'CBENEF_GO_x_CFOP.xlsx'),
    # 'SP': os.path.join(_DATA_DIR, 'CBENEF_SP_x_CFOP.xlsx'),
    # 'MG': os.path.join(_DATA_DIR, 'CBENEF_MG_x_CFOP.xlsx'),
}

# ============================================================
# TABELA NCM (opcional) — usada na aba REGRA VALOR DECLARATÓRIO
# ============================================================
# Tabela oficial NCM vigente, baixada de https://www.gov.br/siscomex/
# Quando setada, a aba REGRA VALOR DECLARATÓRIO traz uma coluna extra
# 'NCMs DA SFT' mostrando os NCMs reais da SFT para cada combinação
# CST+CFOP+FILIAL, com descrição curta — facilita avaliar se o CBENEF
# faz sentido para o seu negócio.
# Se o arquivo não existir, a coluna fica com apenas os códigos NCM.
NCM_TABLE_FILE = os.path.join(_DATA_DIR, 'Tabela_NCM_Vigente.xlsx')

# Caracteres máximos por descrição NCM exibida (truncamento)
NCM_DESC_MAX = 80

# CSTs que admitem benefício fiscal (sugestão CBENEF só ativa nesses)
# 20=Redução, 30=Isento ST, 40=Isenta, 41=Não tributada,
# 50=Suspensão, 70=Com redução e ST
CST_COM_BENEFICIO = {20, 30, 40, 41, 50, 70}

# ============================================================
# CONFIGURAÇÃO DAS RETENÇÕES
# ============================================================
# Cada retenção define: sigla (2 letras), nome, e as 3 colunas físicas da SFT
# (base, alíquota, valor). Para ISS Retido, a base/alíq/valor vem do ICMS
# e o filtro é o flag 'ISS Retido' == 1, restrito a CFOPs de serviço.
#
# Tupla: (sigla, nome, col_base, col_aliq, col_valor, col_flag, val_flag, cfop_allow)
#   col_flag/val_flag: se preenchidos, o filtro de retenção usa o flag em vez de base>0
#   cfop_allow:        set de CFOPs aceitos; None = sem restrição
RETENCOES_CONFIG = [
    ('PR',   'PIS Retido',     'Base Pis',     'Aliq. PIS.1',  'Valor Pis',      None,         None, None),
    ('CR',   'COFINS Retido',  'Base Cofins',  'Aliq. Cofins', 'Valor Cofins',   None,         None, None),
    ('SR',   'CSLL Retido',    'Base CSLL.1',  'Aliq. CSLL.1', 'Valor CSLL.1',   None,         None, None),
    ('IR',   'IRRF',           'Base IRRF',    'Aliq. IRRF',   'Valor IRRF',     None,         None, None),
    ('NR',   'INSS',           'Base INSS',    'Aliq. INSS',   'Valor INSS',     None,         None, None),
    ('TR',   'ISS Retido',     'Base ICMS',    'Alíq. ICMS',   'Valor ICMS',     'ISS Retido', 1,    CFOP_ISS),
]

# ============================================================
# MAPEAMENTO CFOP (estende a tabela oficial)
# ============================================================
EXTRA_CFOP = {
    '1152':'Transferência para industrialização',
    '1353':'Aquisição de mercadoria com substituição tributária',
    '1551':'Compra de bem para o ativo imobilizado',
    '1912':'Retorno de remessa para industrialização',
    '1913':'Retorno de remessa de insumos do exterior',
    '1914':'Retorno de remessa para venda fora do estabelecimento',
    '1915':'Entrada - energia elétrica',
    '1933':'Entrada para uso e consumo',
    '2303':'Aquisição de serviço de comunicação',
    '2353':'Aquisição de mercadoria com ST (interestadual)',
    '2551':'Compra de bem para ativo imobilizado (interestadual)',
    '2910':'Retorno de bonificação/brinde',
    '2912':'Entrada em demonstração',
    '2916':'Retorno de armazém geral',
    '2917':'Retorno de depósito fechado',
    '2923':'Entrada em armazém geral/depósito fechado',
    '2933':'Entrada para uso e consumo (interestadual)',
    '2949':'Outra entrada não especificada (interestadual)',
    '5102':'Venda de mercadoria adquirida ou recebida de terceiros',
    '5152':'Transferência de mercadoria',
    '5202':'Devolução de compra para industrialização',
    '5405':'Venda com retenção de ST',
    '5411':'Devolução de compra com ST',
    '5556':'Venda de bem do ativo imobilizado',
    '5910':'Remessa em bonificação/brinde',
    '5912':'Remessa de amostra grátis',
    '5913':'Remessa para demonstração',
    '5914':'Remessa para venda fora do estabelecimento',
    '5915':'Remessa para armazém geral',
    '5916':'Remessa para depósito fechado',
    '5923':'Remessa para industrialização sob encomenda',
    '5933':'Prestação de serviço - ISSQN',
    '6102':'Venda de mercadoria adquirida/recebida de terceiros (interestadual)',
    '6108':'Venda para Zona Franca de Manaus',
    '6119':'Venda em bonificação (interestadual)',
    '6202':'Devolução de compra para industrialização (interestadual)',
    '6404':'Venda com retenção de ST (interestadual)',
    '6910':'Remessa em bonificação/brinde (interestadual)',
    '6912':'Remessa de amostra grátis (interestadual)',
    '6913':'Remessa para demonstração (interestadual)',
    '6915':'Remessa para armazém geral (interestadual)',
    '6933':'Prestação de serviço - ISSQN (interestadual)',
    '6949':'Outra saída não especificada (interestadual)',
}

ORIGEM_DESC = {
    0: 'Nacional',
    1: 'Estrangeira - Importação direta',
    2: 'Estrangeira - Adquirida mercado interno',
    3: 'Nacional - Conteúdo Importação > 40%',
    4: 'Nacional - produção conforme processos básicos',
    5: 'Nacional - Conteúdo Importação <= 40%',
    6: 'Estrangeira - Importação direta sem similar nacional',
    7: 'Estrangeira - Adquirida mercado interno sem similar nacional',
    8: 'Nacional - Conteúdo Importação > 70%',
}

CST_PC_DESC = {
    1:  'Operação Tributável - Alíquota Básica',
    4:  'Operação Tributável - Alíquota Zero',
    8:  'Operação sem Incidência da Contribuição',
    49: 'Outras Operações de Saída',
    70: 'Operações de Aquisição - Mercado Interno',
    98: 'Outras Operações de Entrada',
}

# ============================================================
# PADRÃO DE CÓDIGOS: [Uso][Imposto][Sequencial 3 dígitos]
# ICMS=IC  PIS-COFINS=PC
# B=Base A=Alíquota E=Escrituração L=Local P=Produto O=Operação F=Fornecedor C=Cliente
# ============================================================

def gen_ids(uso, imposto, n, start=1, pad=3):
    """Gera lista de IDs: ex. gen_ids('B','IC',3) → ['BIC001','BIC002','BIC003']"""
    return [f'{uso}{imposto}{str(i).zfill(pad)}' for i in range(start, start + n)]


# ============================================================
# HELPERS
# ============================================================

_FILIAL_NOT_MAPPED_WARNED = set()  # evita repetir warnings

def normalize_filial(f):
    """
    Normaliza o valor da filial para string consistente.

    Aceita: NaN/None → ''
            float    → '10101' (sem .0)
            int      → '10101'
            string   → str.strip()
    """
    if f is None:
        return ''
    try:
        if pd.isna(f):
            return ''
    except (TypeError, ValueError):
        pass
    if isinstance(f, float):
        if f.is_integer():
            return str(int(f))
        return str(f)
    if isinstance(f, int):
        return str(f)
    return str(f).strip()


def filial_to_state(f):
    """
    Resolve estado da filial a partir do dicionário FILIAL_STATE.

    Tenta múltiplas formas da chave: como veio, normalizada, sem zeros à
    esquerda. Se nada bater, devolve 'XX' e avisa uma única vez por filial.
    """
    raw  = f
    norm = normalize_filial(f)

    # tentativas: forma normalizada, sem zeros à esquerda, vazio, número puro
    tentativas = [norm]
    if norm.lstrip('0') and norm.lstrip('0') != norm:
        tentativas.append(norm.lstrip('0'))
    if norm == '':
        tentativas.append('')
    try:
        tentativas.append(str(int(float(norm))))
    except (ValueError, TypeError):
        pass

    for t in tentativas:
        if t in FILIAL_STATE:
            return FILIAL_STATE[t]

    # Não mapeada
    if norm not in _FILIAL_NOT_MAPPED_WARNED:
        _FILIAL_NOT_MAPPED_WARNED.add(norm)
        print(f"  ⚠ Filial não mapeada em FILIAL_STATE: {raw!r} (normalizado: {norm!r}). "
              f"Adicione no dicionário. Usando 'XX' temporariamente.")
    return 'XX'


def resolve_col(df, logical_name):
    """
    Retorna o nome real da coluna no df que corresponde ao nome lógico
    (busca em COL_ALIASES). Se nenhum alias existir, retorna None.
    """
    aliases = COL_ALIASES.get(logical_name, [logical_name])
    for alias in aliases:
        if alias in df.columns:
            return alias
    return None


def detect_sheet(file_path):
    """
    Detecta automaticamente a aba da SFT.
    Preferência: nomes conhecidos > primeira aba.
    """
    xl = pd.ExcelFile(file_path)
    names = xl.sheet_names
    if len(names) == 1:
        return names[0]
    # Procurar por nomes conhecidos
    for known in ['BASE', 'scjqpw20', 'SFT', 'Plan1', 'Sheet1']:
        if known in names:
            return known
    # Default: primeira
    return names[0]


def safe_float(v):
    try: return float(v) if not pd.isna(v) else 0.0
    except: return 0.0

def safe_int(v):
    try: return int(float(v)) if not pd.isna(v) else 0
    except: return 0

def safe_str(v):
    try:
        if pd.isna(v): return ''
        f = float(v)
        return str(int(f)) if f == int(f) else str(v)
    except:
        return str(v).strip()

def parse_cst_icms(v):
    """
    Separa o CST ICMS de 3 dígitos em (origem_produto, sit_trib_2d).
    Exemplos: 0 → (0,0) | 20 → (0,20) | 200 → (2,0) | 220 → (2,20) | 741 → (7,41)
    """
    s = str(int(float(v))).zfill(3)
    return int(s[0]), int(s[1:])

def calc_base_formula(base_calc, base_icms, sit_trib_2d, aliq_icms):
    """
    Calcula a fórmula de base ICMS para o item.
    base_calc = Vlr Total - Vlr Desconto + Valor Frete + Valor IPI + Vlr Seguro + Vlr Despesas
    
    Retorna: "ISENTO / NÃO TRIBUTADO" | "SEM REDUÇÃO" | "REDUÇÃO DE X%"
    """
    # CSTs intrinsecamente isentos
    ISENTOS_SIT = {41, 60}

    bc = round(float(base_calc), 2)
    bi = round(float(base_icms), 2)

    if sit_trib_2d in ISENTOS_SIT:
        return 'ISENTO / NÃO TRIBUTADO'
    if aliq_icms == 0 and bi == 0:
        return 'ISENTO / NÃO TRIBUTADO'
    if bc <= 0:
        return 'SEM REDUÇÃO'
    if abs(bc - bi) < 0.02:
        return 'SEM REDUÇÃO'
    if bi > bc:  # anomalia: base maior que valor calculado
        return 'SEM REDUÇÃO'

    pct = round((1 - bi / bc) * 100, 2)
    if pct <= 0:
        return 'SEM REDUÇÃO'
    return f'REDUÇÃO DE {pct}%'


def validar_icms(sit_trib_2d, aliq_icms, base_formula_canonica):
    """
    Valida consistência entre CST ICMS, alíquota e fórmula de base.
    IMPORTANTE: usa base_formula_canonica (mode por SIT+ALIQ), não campos brutos.
    Retorna string de observação se houver inconsistência, '' se OK.
    """
    obs = []
    isento = base_formula_canonica == 'ISENTO / NÃO TRIBUTADO'
    sem    = base_formula_canonica == 'SEM REDUÇÃO'
    red    = base_formula_canonica.startswith('REDUÇÃO')

    if sit_trib_2d == 0:   # Tributada integralmente
        if aliq_icms == 0: obs.append("CST 00 exige alíquota > 0 (tributação integral)")
        if isento:         obs.append("CST 00 não deve ter base zerada (tributação integral)")
        if red:            obs.append("CST 00 não deve ter redução — verificar se deveria ser CST 20")

    elif sit_trib_2d == 20:  # Com redução de base
        if aliq_icms == 0: obs.append("CST 20 exige alíquota > 0")
        if isento:         obs.append("CST 20 não deve ter base zerada — verificar se deveria ser CST 40/41")
        if sem:            obs.append("CST 20 sem redução detectada — verificar se deveria ser CST 00")

    elif sit_trib_2d == 41:  # Não tributada
        if aliq_icms > 0:  obs.append("CST 41 não deve ter alíquota > 0")
        if not isento:     obs.append("CST 41 não deve ter base de cálculo")

    elif sit_trib_2d == 60:  # ST cobrado anteriormente
        if aliq_icms > 0:  obs.append("CST 60 não deve ter alíquota própria > 0")
        if not isento:     obs.append("CST 60 não deve ter base de cálculo própria")

    elif sit_trib_2d == 90:  # Outras
        if aliq_icms > 0 and sem: obs.append("CST 90 com alíquota > 0 e base integral — verificar se deveria ser CST 00")
        if aliq_icms > 0 and red: obs.append("CST 90 com alíquota > 0 e redução — verificar se deveria ser CST 20")

    return " | ".join(obs)


def dedup_perfis(df_grouped, num_col, dim_cols, codigo_col, sigla_prefix, padding=3,
                 scope_cols=None):
    """
    Deduplica perfis pelo CONJUNTO de dimensões.

    Cenários com o mesmo conjunto agrupado (ex.: mesmo conjunto de UFs)
    compartilham o mesmo código de perfil.

    Args:
      df_grouped: DataFrame já com [num_col, *dim_cols], 1 linha por (cenário, item).
      num_col:    coluna que identifica o cenário (ex.: 'NUM')
      dim_cols:   lista de colunas que definem o conjunto
      codigo_col: nome da nova coluna que vai receber o código deduplicado
      sigla_prefix: prefixo do código (ex.: 'LIC' → LIC001, LIC002...)
      padding:    zfill do sequencial
      scope_cols: opcional. Colunas que isolam o escopo da dedup (ex.:
                  ['ESTADO_FILIAL'] faz a dedup acontecer DENTRO de cada filial).
                  Cenários com mesma assinatura mas escopos diferentes ainda
                  recebem códigos diferentes (sequencial reinicia por escopo).

    Retorna:
      df_grouped: com a coluna codigo_col preenchida
      num_map:    dict {num_cenario: codigo_deduplicado}
    """
    scope_cols = scope_cols or []
    # A assinatura inclui as scope_cols, garantindo separação entre escopos
    effective_dims = list(scope_cols) + list(dim_cols)

    sig = (
        df_grouped.sort_values([num_col] + effective_dims)
        .groupby(num_col)[effective_dims]
        .apply(lambda g: tuple(sorted(map(tuple, g.values.tolist()))))
        .reset_index(name='_SIG')
    )

    sig_to_code = {}
    seq = 0
    nums_to_code = {}
    for _, row in sig.sort_values(num_col).iterrows():
        s = row['_SIG']
        if s not in sig_to_code:
            seq += 1
            sig_to_code[s] = f"{sigla_prefix}{str(seq).zfill(padding)}"
        nums_to_code[row[num_col]] = sig_to_code[s]

    df_grouped = df_grouped.copy()
    df_grouped[codigo_col] = df_grouped[num_col].map(nums_to_code)
    return df_grouped, nums_to_code


def dedup_perfis_participante(pt_df_in, num_col, sigla_imp, padding=3):
    """
    Deduplica perfil participante, com prefixo F/C conforme TIPO_PART.

    pt_df_in deve conter: [num_col, 'TIPO_PART', 'CLI_FORN']
    Retorna (pt_df com coluna codigo, dict {(num, tipo_part): codigo})
    """
    pt_df = pt_df_in.copy().sort_values([num_col,'TIPO_PART','CLI_FORN'])
    sig = (
        pt_df.groupby([num_col,'TIPO_PART'])['CLI_FORN']
        .apply(lambda g: tuple(sorted(g.tolist())))
        .reset_index(name='_SIG')
    )
    sig_to_code = {}
    seq = 0
    nums_part_to_code = {}
    for _, row in sig.sort_values(num_col).iterrows():
        key = (row['TIPO_PART'], row['_SIG'])
        if key not in sig_to_code:
            seq += 1
            prefix = 'F' if row['TIPO_PART']=='FORNECEDOR' else 'C'
            sig_to_code[key] = f"{prefix}{sigla_imp}{str(seq).zfill(padding)}"
        nums_part_to_code[(row[num_col], row['TIPO_PART'])] = sig_to_code[key]
    pt_df['CIC'] = pt_df.apply(lambda r: nums_part_to_code[(r[num_col], r['TIPO_PART'])], axis=1)
    return pt_df, nums_part_to_code


def validar_pis_cofins(cst_pis, aliq_pis, cst_cof, aliq_cof):
    """
    Valida consistência entre CST e alíquotas de PIS/COFINS.
    """
    obs = []
    if cst_pis == 1 and aliq_pis == 0:    obs.append("CST PIS 01 (Tributável) exige alíquota > 0")
    if cst_cof == 1 and aliq_cof == 0:    obs.append("CST COF 01 (Tributável) exige alíquota > 0")
    if cst_pis == 4 and aliq_pis > 0:     obs.append("CST PIS 04 (Alíq Zero) não deve ter alíquota > 0")
    if cst_cof == 4 and aliq_cof > 0:     obs.append("CST COF 04 (Alíq Zero) não deve ter alíquota > 0")
    if cst_pis in (8, 49) and aliq_pis > 0: obs.append(f"CST PIS {str(cst_pis).zfill(2)} não deve ter alíquota > 0")
    if cst_cof in (8, 49) and aliq_cof > 0: obs.append(f"CST COF {str(cst_cof).zfill(2)} não deve ter alíquota > 0")
    if cst_pis == 70 and aliq_pis > 0:    obs.append("CST PIS 70 com alíquota > 0 — verificar enquadramento")
    if cst_cof == 70 and aliq_cof > 0:    obs.append("CST COF 70 com alíquota > 0 — verificar enquadramento")
    return " | ".join(obs)


# ============================================================
# CARREGAMENTO DOS DADOS
# ============================================================

def _parse_csts_cbenef(s):
    """Parseia string de CSTs do CBENEF em set de ints.
    Exemplos: '30,40,41' → {30,40,41} | '30.41' (Excel) → {30,41} | '20,70' → {20,70}
    """
    s = str(s).replace('.', ',')   # Excel transformou '30,41' em '30.41' (float)
    out = set()
    for p in s.split(','):
        p = p.strip()
        try:
            out.add(int(p))
        except ValueError:
            pass
    return out


def _parse_cfops_cbenef(s):
    """Extrai CFOPs de uma string. Retorna set de strings de 4 dígitos."""
    return set(re.findall(r'\b\d{4}\b', str(s)))


def load_ncm_table(path):
    """
    Carrega a tabela NCM vigente. Retorna dict {NCM_clean: descrição}.

    A tabela oficial tem cabeçalho na linha 5 (header=4 zero-indexed) e
    NCMs com pontos (ex.: '0101.21.00'). Esta função normaliza removendo
    pontos.

    Se o arquivo não existir, retorna dict vazio (script continua).
    """
    try:
        df = pd.read_excel(path, sheet_name='Tabela NCM', header=4)
    except (FileNotFoundError, ValueError):
        # Tentar sem especificar sheet
        try:
            df = pd.read_excel(path, header=4)
        except FileNotFoundError:
            print(f"  ⚠ Tabela NCM não encontrada ({path}). Aba sem descrição NCM.")
            return {}
        except Exception as e:
            print(f"  ⚠ Erro ao ler tabela NCM: {e}. Aba sem descrição NCM.")
            return {}

    if 'Código' not in df.columns or 'Descrição' not in df.columns:
        print(f"  ⚠ Tabela NCM com colunas inesperadas ({list(df.columns)[:3]}...). "
              f"Aba sem descrição NCM.")
        return {}

    df['NCM_clean'] = df['Código'].astype(str).str.strip().str.replace('.', '', regex=False)
    df['Descrição'] = df['Descrição'].astype(str).str.strip()
    mapping = dict(zip(df['NCM_clean'], df['Descrição']))
    print(f"  Tabela NCM: {len(mapping)} códigos carregados")
    return mapping


def desc_ncm(ncm_clean, ncm_map, max_len=NCM_DESC_MAX):
    """
    Retorna descrição curta para um NCM, navegando pela hierarquia
    (8d → 6d → 4d → 2d) para garantir contexto útil.

    Concatena nível mais geral + mais específico, evitando "Outros" sozinho.
    """
    if not ncm_map or not ncm_clean or pd.isna(ncm_clean):
        return ''
    ncm = str(ncm_clean).strip()
    if not ncm:
        return ''

    partes = []
    for n in [ncm[:2], ncm[:4], ncm[:6], ncm]:
        if n and n in ncm_map:
            d = ncm_map[n].strip(' .-')
            if d and d.lower() != 'outros' and d not in partes:
                partes.append(d)
    if not partes:
        return ''
    full = ' / '.join(partes)
    if len(full) > max_len:
        full = full[:max_len-3] + '...'
    return full


def load_cbenef_tables(tables_config):
    """
    Carrega as tabelas CBENEF de cada estado configurado em CBENEF_TABLES.

    Retorna:
      dict {UF: DataFrame com colunas
            ['CBENEF', 'CSTs', 'TIPO_BENEF', 'CFOPs', 'DESC_CBENEF']}
    """
    out = {}
    for uf, path in tables_config.items():
        try:
            df = pd.read_excel(path, sheet_name='CBENEF x CFOP')
        except FileNotFoundError:
            print(f"  ⚠ CBENEF para {uf}: arquivo não encontrado ({path}). Pulando.")
            continue
        except ValueError:
            # aba não existe — tentar primeira
            df = pd.read_excel(path)
        # Normalizar
        df_norm = pd.DataFrame({
            'CBENEF':       df['CÓDIGO CBENEF'].astype(str).str.strip(),
            'CSTs':         df['CST'].apply(_parse_csts_cbenef),
            'TIPO_BENEF':   df['TIPO BENEFÍCIO'].astype(str).str.strip(),
            'CFOPs':        df['CFOP(s) INDICADO(s)'].apply(_parse_cfops_cbenef),
            'DESC_CBENEF':  df['DESCRIÇÃO CBENEF (NFA-e)'].astype(str),
        })
        out[uf] = df_norm
        print(f"  CBENEF/{uf}: {len(df_norm)} códigos carregados")
    return out


def sugerir_cbenef(cbenef_tables, uf_filial, sit_trib_2d, cfops_set, base_formula):
    """
    Sugere CBENEF(s) para um cenário ICMS.

    Args:
      cbenef_tables: dict {UF: DataFrame normalizado}
      uf_filial:     estado da filial (ex.: 'GO')
      sit_trib_2d:   CST sem origem do produto (2 dígitos, ex.: 41)
      cfops_set:     set de CFOPs do perfil operação do cenário
      base_formula:  string BASE_FORMULA do cenário (para filtrar tipo de benefício)

    Retorna:
      string com CBENEFs separados por vírgula, ou
      "Verificar manualmente" se cenário tem benefício mas nada bate, ou
      '' se cenário não admite benefício.
    """
    # Só sugere se o CST estiver na lista de benefícios
    if sit_trib_2d not in CST_COM_BENEFICIO:
        return ''

    tbl = cbenef_tables.get(uf_filial)
    if tbl is None or len(tbl) == 0:
        return ''  # estado sem tabela cadastrada

    # Filtrar por CST
    mask_cst  = tbl['CSTs'].apply(lambda s: sit_trib_2d in s)
    # Filtrar por intersecção de CFOPs (qualquer CFOP do cenário em qualquer CFOP da regra)
    mask_cfop = tbl['CFOPs'].apply(lambda s: bool(s & cfops_set))

    candidatos = tbl[mask_cst & mask_cfop]

    if len(candidatos) == 0:
        return 'Verificar manualmente'

    # Opcional: filtrar por tipo de benefício compatível com BASE_FORMULA
    # ISENTO/NÃO TRIB → Isenção, Não Incidência | REDUÇÃO → Redução BC
    bf = str(base_formula).upper()
    if bf.startswith('REDUÇÃO') or bf.startswith('REDUCAO'):
        filtrados = candidatos[candidatos['TIPO_BENEF'].str.contains('Redução', case=False, na=False)]
        if len(filtrados) > 0:
            candidatos = filtrados
    elif 'ISENTO' in bf or 'NÃO TRIB' in bf or 'NAO TRIB' in bf:
        filtrados = candidatos[candidatos['TIPO_BENEF'].str.contains(
            'Isenção|Não Incidência|Isen|Imune', case=False, na=False, regex=True)]
        if len(filtrados) > 0:
            candidatos = filtrados

    return ', '.join(candidatos['CBENEF'].tolist())


def coletar_perfil_empresa(perguntas):
    """
    Prompt interativo no terminal para coletar perfil da empresa.

    Cada pergunta tem default (Enter aceita). Resposta:
      's', 'sim', 'y', 'yes', '1', '' (se default='s') → True
      'n', 'nao', 'não', 'no', '0', '' (se default='n') → False

    Retorna dict {flag: bool}.
    """
    perfil = {}
    if not perguntas:
        return perfil

    print()
    print("=" * 70)
    print(" QUESTIONÁRIO DE PERFIL DA EMPRESA")
    print(" Responda 's' (sim) ou 'n' (não). Enter = resposta padrão.")
    print(" Suas respostas filtram CBENEFs que não se aplicam ao seu negócio.")
    print("=" * 70)
    print()

    for flag, pergunta, default in perguntas:
        default_label = '[S/n]' if default == 's' else '[s/N]'
        while True:
            try:
                resp = input(f"  {pergunta} {default_label}: ").strip().lower()
            except EOFError:
                # Sem stdin (ex.: rodando em pipe) — usa default
                resp = ''
            if resp == '':
                resp = default
            if resp in ('s', 'sim', 'y', 'yes', '1'):
                perfil[flag] = True
                break
            elif resp in ('n', 'nao', 'não', 'no', '0'):
                perfil[flag] = False
                break
            else:
                print(f"    Resposta inválida '{resp}'. Use 's' ou 'n'.")

    print()
    print("=" * 70)
    n_sim = sum(1 for v in perfil.values() if v)
    n_nao = sum(1 for v in perfil.values() if not v)
    print(f" Perfil capturado: {n_sim} características SIM, {n_nao} NÃO")
    print("=" * 70)
    print()
    return perfil


def _avaliar_aplicavel(classif_entry, perfil_empresa, ncms_sft):
    """
    Avalia se um CBENEF é aplicável dado o perfil da empresa e os NCMs da SFT.

    Retorna tupla (aplicavel, motivo_descarte):
      aplicavel: 'SIM' | 'TALVEZ' | 'NÃO'
      motivo_descarte: string explicando o porquê (vazio se SIM)
    """
    if not classif_entry:
        return ('TALVEZ', 'Código não classificado')

    cat = classif_entry['cat']
    if cat == 'IGNORE':
        return ('NÃO', 'Reconhecida judicialmente — não aplica em cadastro padrão')

    flags = classif_entry.get('flags', []) or []

    # Se TEM flags e usuário disse NÃO para TODAS → descarta com motivo
    if flags:
        flags_que_disse_sim = [f for f in flags if perfil_empresa.get(f) is True]
        flags_que_disse_nao = [f for f in flags if perfil_empresa.get(f) is False]
        if flags and not flags_que_disse_sim and len(flags_que_disse_nao) == len(flags):
            # Todas as flags relevantes foram negadas pelo perfil
            return ('NÃO', f"Perfil indica que não se aplica ({', '.join(flags)})")

    # PRODUTO: precisa que NCM bate
    if cat == 'PRODUTO':
        chapters = classif_entry.get('ncm_chapters') or []
        bate = any(_ncm_match_chapter(n, chapters) for n in ncms_sft)
        if not bate:
            return ('NÃO', f"NCM da SFT não bate com produto descrito (capítulos esperados: {','.join(chapters)})")
        # NCM bate. Se também há flags e perfil confirma, é SIM. Se há flags sem confirmação, TALVEZ.
        if flags and any(perfil_empresa.get(f) is True for f in flags):
            return ('SIM', '')
        elif flags:
            return ('TALVEZ', f"NCM bate, mas verifique se aplica ao contexto ({', '.join(flags)})")
        return ('SIM', '')

    # GENERICO: sempre SIM (se passou pelos filtros de flag acima)
    if cat == 'GENERICO':
        if flags and any(perfil_empresa.get(f) is True for f in flags):
            return ('SIM', '')
        elif flags:
            return ('TALVEZ', f"Verifique se aplica ({', '.join(flags)})")
        return ('SIM', '')

    # CONTEXTO: TALVEZ (revisar manualmente) se passou pelos filtros de flag
    if cat == 'CONTEXTO':
        if flags and any(perfil_empresa.get(f) is True for f in flags):
            return ('TALVEZ', f"Contexto específico — confirme detalhes")
        elif flags:
            # Sem confirmação positiva mas com flags todas negadas já foi descartado acima
            return ('TALVEZ', f"Contexto específico — revisar")
        return ('TALVEZ', 'Contexto específico — revisar manualmente')

    return ('TALVEZ', '')


def _ncm_match_chapter(ncm, prefixos):
    """Verifica se um NCM (8 dígitos) começa com algum dos prefixos."""
    if not ncm or not prefixos: return False
    ncm = str(ncm)
    return any(ncm.startswith(p) for p in prefixos)


def build_valor_declaratorio_table(df, cbenef_tables, ncm_map=None,
                                    classificacao_por_estado=None,
                                    perfil_empresa=None):
    """
    Monta a aba REGRA VALOR DECLARATÓRIO com filtragem dupla:
      1) Por categoria/NCM (PRODUTO precisa bater com NCM da SFT)
      2) Por perfil da empresa (flags incompatíveis → NÃO)

    Mantém TODOS os candidatos na planilha, marcados na coluna APLICÁVEL?:
      ✅ SIM    - candidato forte (GENÉRICO com flag positiva, ou PRODUTO compatível confirmado)
      ⚠ TALVEZ - revisar manualmente (CONTEXTO, ou produto sem confirmação de flag)
      ❌ NÃO   - perfil ou NCM indica que não aplica

    Args:
      df:            DataFrame processado da SFT
      cbenef_tables: dict {UF: DataFrame normalizado} de CBENEF
      ncm_map:       dict {NCM_clean: descrição} (opcional)
      classificacao_por_estado: dict {UF: {CBENEF_CODE: classif_entry}}
      perfil_empresa: dict {flag: bool} do questionário (None = só usa NCM)

    Retorna:
      DataFrame com 10 colunas: + APLICÁVEL? + OBSERVAÇÃO
    """
    cols_out = ['TRIBUTO','FILIAL','CÓDIGO VALOR DECLARATÓRIO','CST','CFOP',
                'TIPO BENEFÍCIO','DESCRIÇÃO CBENEF','NCMs DA SFT',
                'APLICÁVEL?','OBSERVAÇÃO']
    out_rows = []
    classificacao_por_estado = classificacao_por_estado or {}
    perfil_empresa = perfil_empresa or {}

    if not cbenef_tables:
        return pd.DataFrame(out_rows, columns=cols_out)

    # Filtrar SFT: só ICMS, só CST com benefício
    d = df[df['IMPOSTO'] == 'ICMS'].copy()
    d = d[d['SIT_TRIB'].isin(CST_COM_BENEFICIO)].copy()
    if len(d) == 0:
        return pd.DataFrame(out_rows, columns=cols_out)

    # NCMs por (FILIAL, SIT_TRIB, CFOP)
    ncms_por_combo = (
        d.groupby(['FILIAL', 'SIT_TRIB', 'CFOP'])['NCM']
        .apply(lambda s: sorted(set(str(x) for x in s if x and str(x) != 'nan')))
        .to_dict()
    )

    combos = (
        d.groupby(['ESTADO_FILIAL', 'FILIAL', 'SIT_TRIB', 'CFOP', 'BASE_FORMULA'])
        .size().reset_index(name='_qtd')
    )

    seen = set()
    stats = {'SIM': 0, 'TALVEZ': 0, 'NAO': 0}

    for _, c in combos.iterrows():
        uf_filial   = c['ESTADO_FILIAL']
        filial      = c['FILIAL']
        sit_trib    = int(c['SIT_TRIB'])
        cfop        = str(c['CFOP'])
        base_form   = str(c['BASE_FORMULA'])

        tbl = cbenef_tables.get(uf_filial)
        if tbl is None or len(tbl) == 0:
            continue

        classif = classificacao_por_estado.get(uf_filial, {})

        mask_cst  = tbl['CSTs'].apply(lambda s: sit_trib in s)
        mask_cfop = tbl['CFOPs'].apply(lambda s: cfop in s)
        cand = tbl[mask_cst & mask_cfop]

        if len(cand) == 0:
            continue

        # Filtra por tipo de benefício coerente com BASE_FORMULA
        bf = base_form.upper()
        if bf.startswith('REDUÇÃO') or bf.startswith('REDUCAO'):
            filt = cand[cand['TIPO_BENEF'].str.contains('Redução', case=False, na=False)]
            if len(filt) > 0: cand = filt
        elif 'ISENTO' in bf or 'NÃO TRIB' in bf or 'NAO TRIB' in bf:
            filt = cand[cand['TIPO_BENEF'].str.contains(
                'Isenção|Não Incidência|Imune', case=False, na=False, regex=True)]
            if len(filt) > 0: cand = filt

        # NCMs reais
        ncms = ncms_por_combo.get((filial, sit_trib, cfop), [])
        if ncm_map:
            ncm_str = '\n'.join(
                f"{n}  —  {desc_ncm(n, ncm_map)}" if desc_ncm(n, ncm_map) else n
                for n in ncms
            )
        else:
            ncm_str = '\n'.join(ncms)

        for _, k in cand.iterrows():
            cbenef_code = k['CBENEF']
            key = (str(filial), cbenef_code, sit_trib, cfop)
            if key in seen:
                continue

            classif_entry = classif.get(cbenef_code)

            # Avaliar aplicabilidade
            aplicavel, motivo = _avaliar_aplicavel(classif_entry, perfil_empresa, ncms)

            # IGNORE: descarta totalmente (não vai pra planilha)
            if classif_entry and classif_entry.get('cat') == 'IGNORE':
                continue

            # Stats
            if aplicavel == 'SIM':
                stats['SIM'] += 1
            elif aplicavel == 'TALVEZ':
                stats['TALVEZ'] += 1
            else:
                stats['NAO'] += 1

            # Compor observação final
            nota_classif = (classif_entry or {}).get('nota', '')
            if aplicavel == 'NÃO':
                observacao = f"❌ {motivo}"
            elif aplicavel == 'TALVEZ':
                if motivo:
                    observacao = f"⚠ {motivo} — {nota_classif}"
                else:
                    observacao = f"⚠ {nota_classif}"
            else:
                observacao = f"✅ {nota_classif}"

            seen.add(key)
            out_rows.append({
                'TRIBUTO':                    'ICMS',
                'FILIAL':                     filial,
                'CÓDIGO VALOR DECLARATÓRIO':  cbenef_code,
                'CST':                        str(sit_trib).zfill(2),
                'CFOP':                       cfop,
                'TIPO BENEFÍCIO':             k['TIPO_BENEF'],
                'DESCRIÇÃO CBENEF':           k['DESC_CBENEF'],
                'NCMs DA SFT':                ncm_str,
                'APLICÁVEL?':                 aplicavel,
                'OBSERVAÇÃO':                 observacao,
            })

    # Print stats
    print(f"  Filtragem CBENEF:")
    print(f"    ✅ SIM      (candidatos fortes):  {stats['SIM']:4d}")
    print(f"    ⚠ TALVEZ   (revisar manualmente):{stats['TALVEZ']:4d}")
    print(f"    ❌ NÃO      (descartados pelo perfil/NCM): {stats['NAO']:4d}")

    if not out_rows:
        return pd.DataFrame(columns=cols_out)

    result = pd.DataFrame(out_rows)
    # Ordenar: SIM → TALVEZ → NÃO, depois por filial/cst/cfop/código
    ordem_apl = {'SIM': 0, 'TALVEZ': 1, 'NÃO': 2}
    result['_ordem'] = result['APLICÁVEL?'].map(ordem_apl).fillna(9)
    result = result.sort_values(['_ordem','FILIAL','CST','CFOP','CÓDIGO VALOR DECLARATÓRIO']).reset_index(drop=True)
    result = result.drop(columns='_ordem')
    return result


def load_cfop_map(cfop_file):
    """Carrega mapa CFOP código → descrição. Aceita .xls ou .xlsx."""
    cfop_map = {}
    raw = None
    # Tenta primeiro como xlsx (caso o arquivo tenha sido convertido), depois xls
    for engine, ext in [('openpyxl', '.xlsx'), ('xlrd', '.xls')]:
        try:
            path = cfop_file if cfop_file.lower().endswith(ext) else cfop_file.rsplit('.', 1)[0] + ext
            raw = pd.read_excel(path, engine=engine, header=None)
            break
        except Exception:
            continue
    if raw is None:
        # Última tentativa com o nome exato e engine automático
        try:
            raw = pd.read_excel(cfop_file, header=None)
        except Exception as e:
            print(f"Aviso: não foi possível carregar tabela CFOP: {e}")
            cfop_map.update(EXTRA_CFOP)
            return cfop_map

    for _, row in raw.iterrows():
        code = str(row[0]).strip()
        desc = str(row[1]).strip()
        if re.match(r'^\d{4}$', code) and desc not in ('nan', 'NaN', ''):
            cfop_map[code] = desc
    cfop_map.update(EXTRA_CFOP)
    return cfop_map


def load_sft(sft_file, sheet_name):
    """
    Carrega e normaliza colunas numéricas da SFT.

    Se sheet_name for None, detecta automaticamente.
    Se uma coluna não existir, é ignorada (será tratada como 0/vazio depois).
    """
    if sheet_name is None:
        sheet_name = detect_sheet(sft_file)
        print(f"  Sheet detectada: {sheet_name!r}")

    df = pd.read_excel(sft_file, sheet_name=sheet_name)

    # Colunas a converter para numérico — preserva nome real se houver alias
    num_cols_logical = [
        'Vlr Total', 'Vlr Desconto', 'Valor Frete', 'Valor IPI', 'Vlr Seguro', 'Vlr Despesas',
        'Base ICMS', 'Alíq. ICMS', 'Valor ICMS',
        'Base PIS', 'Aliq. PIS', 'Base COFINS', 'Aliq. COFINS',
        'CST Pis', 'CST COF', 'Sit.Tribut.',
        # Retenções
        'Base Pis', 'Valor Pis', 'Aliq. PIS.1',
        'Base Cofins', 'Valor Cofins', 'Aliq. Cofins',
        'Base CSLL.1', 'Valor CSLL.1', 'Aliq. CSLL.1',
        'Base IRRF', 'Aliq. IRRF', 'Valor IRRF',
        'Base INSS', 'Aliq. INSS', 'Valor INSS',
        'ISS Retido',
        # PROTEGE GO
        'Aliq.PROT.GO', 'Bas.PROT.GO', 'Val.PROT.GO',
        # ICMS Complementar
        'Difal ICMS', 'Perc. Destin', 'Base. Destin', 'Vlr FECP Dif',
        # IPI
        'Vlr Base IPI', 'Vlr Isen IPI', 'Vlr Outr IPI', 'Vlr IPI Obs',
        'Alíq IPI', 'Trib. IPI',
    ]
    for logical in num_cols_logical:
        real = resolve_col(df, logical)
        if real is not None:
            df[real] = pd.to_numeric(df[real], errors='coerce').fillna(0)
            # Se o nome real for diferente do lógico, criar coluna espelho
            # para que o resto do código possa usar o nome lógico.
            if real != logical:
                df[logical] = df[real]
    return df


# ============================================================
# PROCESSAMENTO DA SFT → DATAFRAME ENRIQUECIDO
# ============================================================

def process_sft(sft, cfop_map):
    """
    Processa cada linha da SFT e retorna DataFrame com colunas derivadas.

    Marca cada linha como IMPOSTO='ICMS' ou IMPOSTO='ISS' baseado no CFOP.
    Linhas ISS são desviadas do fluxo ICMS no build_icms_tables.
    """
    rows = []
    for _, r in sft.iterrows():
        filial_raw    = normalize_filial(r['Filial'])
        estado_filial = filial_to_state(r['Filial'])
        tipo_mov      = str(r['Tipo Mov.']).strip()
        estado_ref    = str(r['Estado Ref']).strip()
        entrada       = tipo_mov.lower() == 'entrada'

        estado_orig = estado_ref    if entrada else estado_filial
        estado_dest = estado_filial if entrada else estado_ref
        tipo_part   = 'FORNECEDOR' if entrada else 'CLIENTE'
        is_intra    = (estado_orig == estado_dest)

        cfop      = safe_str(r['Cod. Fiscal'])
        cfop_desc = cfop_map.get(cfop, f'CFOP {cfop}')
        aliq_icms = safe_float(r['Alíq. ICMS'])
        cli_forn  = safe_str(r['Cli/Forn.'])
        ncm       = safe_str(r['Cód. NCM'])
        produto   = str(r['Cod. Produto']).strip()

        # Tipo de imposto principal pela CFOP
        imposto = 'ISS' if cfop in CFOP_ISS else 'ICMS'

        cst_raw = safe_int(r['Sit.Tribut.'])
        origem_prod, sit_trib_2d = parse_cst_icms(cst_raw)

        # BASE_CALC = Vlr Total - Vlr Desconto + Frete + IPI + Seguro + Despesas
        base_calc = (
            safe_float(r['Vlr Total']) - safe_float(r['Vlr Desconto'])
            + safe_float(r['Valor Frete']) + safe_float(r['Valor IPI'])
            + safe_float(r['Vlr Seguro']) + safe_float(r['Vlr Despesas'])
        )
        base_formula = calc_base_formula(base_calc, r['Base ICMS'], sit_trib_2d, aliq_icms)

        # PIS/COFINS — colunas AMARELAS (não usar colunas azuis = retenção PCC)
        cst_pis  = safe_int(r['CST Pis'])
        cst_cof  = safe_int(r['CST COF'])
        aliq_pis = safe_float(r['Aliq. PIS'])
        aliq_cof = safe_float(r['Aliq. COFINS'])

        # ISS — campos específicos (sempre lidos; só usados quando IMPOSTO='ISS')
        cod_servico  = safe_str(r['Cód. Serviço'])  if 'Cód. Serviço'  in r else ''
        cod_trib_mun = safe_str(r['C. Trib. Mun']) if 'C. Trib. Mun' in r else ''

        # Retenções — leitura genérica de todas as colunas configuradas
        ret_data = {}
        for sigla, _nome, col_base, col_aliq, col_valor, col_flag, _val_flag, _cfop_allow in RETENCOES_CONFIG:
            ret_data[f'RET_{sigla}_BASE']  = safe_float(r[col_base])  if col_base  in r else 0.0
            ret_data[f'RET_{sigla}_ALIQ']  = safe_float(r[col_aliq])  if col_aliq  in r else 0.0
            ret_data[f'RET_{sigla}_VALOR'] = safe_float(r[col_valor]) if col_valor in r else 0.0
            if col_flag is not None:
                ret_data[f'RET_{sigla}_FLAG'] = safe_int(r[col_flag]) if col_flag in r else 0

        # PROTEGE GO
        aliq_pg = safe_float(r['Aliq.PROT.GO']) if 'Aliq.PROT.GO' in r else 0.0
        base_pg = safe_float(r['Bas.PROT.GO'])  if 'Bas.PROT.GO'  in r else 0.0
        valor_pg = safe_float(r['Val.PROT.GO']) if 'Val.PROT.GO'  in r else 0.0

        # ICMS Complementar — DIFAL clássico e Base Destino
        difal_icms   = safe_float(r['Difal ICMS'])   if 'Difal ICMS'   in r else 0.0
        perc_destin  = safe_float(r['Perc. Destin']) if 'Perc. Destin' in r else 0.0
        base_destin  = safe_float(r['Base. Destin']) if 'Base. Destin' in r else 0.0

        # IPI — usa Trib. IPI como CST
        trib_ipi  = safe_int(r['Trib. IPI'])    if 'Trib. IPI' in r else 0
        aliq_ipi  = safe_float(r['Alíq IPI'])   if 'Alíq IPI'  in r else 0.0
        base_ipi  = safe_float(r['Vlr Base IPI']) if 'Vlr Base IPI' in r else 0.0
        isen_ipi  = safe_float(r['Vlr Isen IPI']) if 'Vlr Isen IPI' in r else 0.0
        outr_ipi  = safe_float(r['Vlr Outr IPI']) if 'Vlr Outr IPI' in r else 0.0
        valor_ipi = safe_float(r['Valor IPI'])    if 'Valor IPI'    in r else 0.0
        # Sinaliza se há registro de IPI (Trib. IPI preenchido)
        tem_ipi = bool(trib_ipi) and 'Trib. IPI' in r and not pd.isna(r['Trib. IPI'])

        rows.append({
            'IMPOSTO':       imposto,
            'FILIAL':        filial_raw,
            'ESTADO_FILIAL': estado_filial,
            'ESTADO_ORIG':   estado_orig,
            'ESTADO_DEST':   estado_dest,
            'IS_INTRA':      is_intra,
            'TIPO_PART':     tipo_part,
            'CLI_FORN':      cli_forn,
            'CFOP':          cfop,
            'CFOP_DESC':     cfop_desc,
            'BASE_FORMULA':  base_formula,
            'ALIQ_ICMS':     aliq_icms,
            'BASE_ICMS':     safe_float(r['Base ICMS']) if 'Base ICMS' in r else 0.0,
            'SIT_TRIB':      sit_trib_2d,
            'ORIGEM_PROD':   origem_prod,
            'NCM':           ncm,
            'PRODUTO':       produto,
            'TIPO_MOV':      tipo_mov,
            'CST_PIS':       cst_pis,
            'CST_COF':       cst_cof,
            'ALIQ_PIS':      aliq_pis,
            'ALIQ_COF':      aliq_cof,
            'COD_SERVICO':   cod_servico,
            'COD_TRIB_MUN':  cod_trib_mun,
            # PROTEGE
            'ALIQ_PG':       aliq_pg,
            'BASE_PG':       base_pg,
            'VALOR_PG':      valor_pg,
            # ICMS Complementar
            'DIFAL_ICMS':    difal_icms,
            'PERC_DESTIN':   perc_destin,
            'BASE_DESTIN':   base_destin,
            # IPI
            'TRIB_IPI':      trib_ipi,
            'ALIQ_IPI':      aliq_ipi,
            'BASE_IPI':      base_ipi,
            'ISEN_IPI':      isen_ipi,
            'OUTR_IPI':      outr_ipi,
            'VALOR_IPI':     valor_ipi,
            'TEM_IPI':       tem_ipi,
            **ret_data,
        })

    df = pd.DataFrame(rows)

    # Canonizar BASE_FORMULA por (IMPOSTO, SIT_TRIB, ALIQ_ICMS) usando mode.
    # IMPOSTO entra no group para não misturar regras de ICMS e ISS, que
    # compartilham os mesmos campos físicos mas são tributos distintos.
    # Elimina variações de arredondamento item a item.
    canon = (
        df.groupby(['IMPOSTO', 'SIT_TRIB', 'ALIQ_ICMS'])['BASE_FORMULA']
        .agg(lambda x: x.value_counts().index[0])
        .reset_index()
        .rename(columns={'BASE_FORMULA': 'BF_CANON'})
    )
    df = df.merge(canon, on=['IMPOSTO', 'SIT_TRIB', 'ALIQ_ICMS'], how='left')
    df['BASE_FORMULA'] = df['BF_CANON'].fillna(df['BASE_FORMULA'])
    df.drop(columns='BF_CANON', inplace=True)

    return df


# ============================================================
# CONSTRUÇÃO DAS TABELAS ICMS
# ============================================================

def build_icms_tables(df, cbenef_tables=None):
    """
    Constrói todas as tabelas de perfis e regras para o ICMS.

    IMPORTANTE: filtra IMPOSTO == 'ICMS' — linhas com CFOPs 1933/2933/5933/6933
    são desviadas para build_iss_tables.

    Args:
      df: DataFrame processado da SFT
      cbenef_tables: mantido para compatibilidade (não usado mais; o cruzamento
                     CBENEF é feito em build_valor_declaratorio_table)

    CHAVE DO CENÁRIO ICMS:
      TIPO_MOV + IS_INTRA + SIT_TRIB + ALIQ_ICMS + BASE_FORMULA
    
    Tudo mais (UF, CFOP, participante, produto, origem) é agrupado DENTRO do cenário.

    Retorna dicionário com todos os DataFrames de saída e o mestre.
    """
    df = df[df['IMPOSTO'] == 'ICMS'].copy().reset_index(drop=True)
    IKEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'IS_INTRA', 'SIT_TRIB', 'ALIQ_ICMS', 'BASE_FORMULA']

    # ── Regra de Base (BIC) ──────────────────────────────────────────────────
    bf_uniq = df['BASE_FORMULA'].drop_duplicates().sort_values().reset_index(drop=True)
    bf_table = pd.DataFrame({'FORMULA': bf_uniq})
    bf_table.insert(0, 'CODIGO DA REGRA', gen_ids('B', 'IC', len(bf_table)))
    bf_table.insert(1, 'TIPO DE BASE', 'ICMS')
    bf_map = {r['FORMULA']: r['CODIGO DA REGRA'] for _, r in bf_table.iterrows()}
    df['BIC'] = df['BASE_FORMULA'].map(bf_map)

    # ── Regra de Alíquota (AIC) ───────────────────────────────────────────────
    aq_table = df[['ALIQ_ICMS']].drop_duplicates().sort_values('ALIQ_ICMS').reset_index(drop=True)
    aq_table.insert(0, 'CODIGO DA REGRA', gen_ids('A', 'IC', len(aq_table)))
    aq_table['ALIQUOTA_DEC'] = aq_table['ALIQ_ICMS'].apply(lambda x: round(x / 100, 4))
    aq_map = {r['ALIQ_ICMS']: r['CODIGO DA REGRA'] for _, r in aq_table.iterrows()}
    df['AIC'] = df['ALIQ_ICMS'].map(aq_map)

    # ── Regra de Escrituração / CST (EIC) ────────────────────────────────────
    esc_table = df[['SIT_TRIB']].drop_duplicates().sort_values('SIT_TRIB').reset_index(drop=True)
    esc_table.insert(0, 'CODIGO DA REGRA', gen_ids('E', 'IC', len(esc_table)))
    esc_table['CST ICMS'] = esc_table['SIT_TRIB'].apply(lambda x: str(x).zfill(2))
    esc_map = {r['SIT_TRIB']: r['CODIGO DA REGRA'] for _, r in esc_table.iterrows()}
    df['EIC'] = df['SIT_TRIB'].map(esc_map)

    # ── Mestre de cenários ICMS ───────────────────────────────────────────────
    mestre = (
        df[IKEY + ['BIC', 'AIC', 'EIC']]
        .drop_duplicates()
        .sort_values(['TIPO_MOV', 'IS_INTRA', 'SIT_TRIB', 'ALIQ_ICMS'])
        .reset_index(drop=True)
    )
    mestre['NUM'] = range(1, len(mestre) + 1)
    mestre['OBS'] = mestre.apply(
        lambda r: validar_icms(r['SIT_TRIB'], r['ALIQ_ICMS'], r['BASE_FORMULA']), axis=1)
    mestre['ERRO'] = mestre['OBS'].str.len() > 0
    mestre['DESCRICAO'] = mestre.apply(lambda r: (
        f"ICMS_{int(r['ALIQ_ICMS'])}%-CST{str(r['SIT_TRIB']).zfill(2)}"
        f"-{'ENT' if r['TIPO_MOV'].lower()=='entrada' else 'SAI'}"
        f"-{'INTRA' if r['IS_INTRA'] else 'INTER'}"
    ), axis=1)

    df = df.merge(mestre[IKEY + ['NUM']], on=IKEY, how='left')

    # ── Perfil Origem/Destino (LIC) — deduplicado por conjunto de UFs ────────
    od_df = (df.groupby(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).size()
               .reset_index(name='_')[['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']]
               .drop_duplicates().sort_values(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).reset_index(drop=True))
    od_df, lic_map = dedup_perfis(od_df, 'NUM',
        ['ESTADO_ORIG', 'ESTADO_DEST'], 'LIC', 'LIC')
    od_out = od_df[['LIC', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates() \
        .sort_values(['LIC','ESTADO_ORIG','ESTADO_DEST']).copy()
    od_out.columns = ['CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO']
    mestre['LIC'] = mestre['NUM'].map(lic_map)

    # ── Perfil Operação (OIC) — deduplicado por conjunto de CFOPs ────────────
    op_df = (df[['NUM', 'CFOP', 'CFOP_DESC', 'TIPO_MOV']].drop_duplicates()
               .sort_values(['NUM', 'CFOP']).reset_index(drop=True))
    op_df, oic_map = dedup_perfis(op_df, 'NUM',
        ['CFOP', 'TIPO_MOV'], 'OIC', 'OIC')
    op_out = op_df[['OIC', 'TIPO_MOV', 'CFOP', 'CFOP_DESC']].drop_duplicates() \
        .sort_values(['OIC','CFOP']).copy()
    op_out.columns = ['CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP']
    mestre['OIC'] = mestre['NUM'].map(oic_map)

    # ── Perfil Participante — deduplicado por conjunto de Cli/Forn ──────────
    pt_df = (df[['NUM', 'TIPO_PART', 'CLI_FORN']].drop_duplicates()
               .sort_values(['NUM', 'TIPO_PART', 'CLI_FORN']).reset_index(drop=True))
    pt_df, cic_map = dedup_perfis_participante(pt_df, 'NUM', 'IC')
    pt_out = pt_df[['CIC', 'TIPO_PART', 'CLI_FORN']].drop_duplicates() \
        .sort_values(['CIC','CLI_FORN']).copy()
    pt_out.columns = ['CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE']
    mestre['CIC'] = mestre.apply(
        lambda r: cic_map.get(
            (r['NUM'], 'FORNECEDOR' if r['TIPO_MOV'].lower()=='entrada' else 'CLIENTE'), ''), axis=1)

    # ── Perfil Produto (PIC) — deduplicado por conjunto de (NCM, Produto, Origem)
    pr_df = (df[['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']].drop_duplicates()
               .sort_values(['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']).reset_index(drop=True))
    pr_df, pic_map = dedup_perfis(pr_df, 'NUM',
        ['NCM', 'PRODUTO', 'ORIGEM_PROD'], 'PIC', 'PIC')
    pr_df['ORIGEM_DESC'] = pr_df['ORIGEM_PROD'].map(ORIGEM_DESC).fillna('Não identificada')
    pr_out = pr_df[['PIC', 'NCM', 'PRODUTO', 'ORIGEM_PROD', 'ORIGEM_DESC']].drop_duplicates() \
        .sort_values(['PIC','NCM','PRODUTO']).copy()
    pr_out.columns = ['CODIGO DA REGRA', 'NCM PRODUTO', 'CODIGO PRODUTO', 'ORIGEM', 'DESCRIÇÃO ORIGEM']
    mestre['PIC'] = mestre['NUM'].map(pic_map)

    mestre['CODIGOS_REGRA'] = mestre.apply(
        lambda r: f"{r['BIC']} | {r['AIC']} | {r['LIC']} | {r['EIC']} | {r['OIC']} | {r['CIC']} | {r['PIC']}", axis=1)

    return {
        'mestre':    mestre,
        'df':        df,
        'bf_table':  bf_table,
        'aq_table':  aq_table,
        'esc_table': esc_table[['CODIGO DA REGRA', 'CST ICMS']],
        'od_out':    od_out,
        'op_out':    op_out,
        'pt_out':    pt_out,
        'pr_out':    pr_out,
    }


# ============================================================
# CONSTRUÇÃO DAS TABELAS ISS
# ============================================================

def validar_iss(aliq_iss, base_formula_canonica):
    """
    Valida consistência ISS (sem CST — só alíquota e base).
    Retorna string de observação se houver inconsistência, '' se OK.
    """
    obs = []
    isento = base_formula_canonica == 'ISENTO / NÃO TRIBUTADO'
    red    = base_formula_canonica.startswith('REDUÇÃO')

    if aliq_iss > 0 and isento:
        obs.append("ISS com alíquota > 0 e base isenta — inconsistência")
    if aliq_iss == 0 and not isento:
        obs.append("ISS com alíquota zero mas base tributada — verificar imunidade/isenção")
    if red and aliq_iss == 0:
        obs.append("ISS com redução de base e alíquota zero — verificar enquadramento")

    return " | ".join(obs)


def build_iss_tables(df):
    """
    Constrói todas as tabelas de perfis e regras para o ISS.

    IMPORTANTE: filtra IMPOSTO == 'ISS' (CFOPs 1933/2933/5933/6933).
    ISS usa as MESMAS colunas físicas do ICMS (Base ICMS, Alíq. ICMS, Valor ICMS).

    CHAVE DO CENÁRIO ISS:
      TIPO_MOV + IS_INTRA + ALIQ_ISS + BASE_FORMULA
    
    ISS NÃO tem CST — diferença principal vs ICMS.
    Perfil Produto/Serviço usa COD_SERVICO + PRODUTO (sem NCM, sem CNAE).
    Origem do produto também não se aplica a serviços.

    Códigos: BIS, AIS, LIS, OIS, FIS/CIS, RIS (Rol de Serviço)

    Retorna dicionário com todos os DataFrames de saída e o mestre.
    """
    df = df[df['IMPOSTO'] == 'ISS'].copy().reset_index(drop=True)

    # Caso não haja linhas ISS na SFT, devolve estrutura vazia para o write_excel pular
    if len(df) == 0:
        empty = pd.DataFrame()
        return {
            'mestre':   empty, 'df': df,
            'bf_table': empty, 'aq_table': empty,
            'od_out':   empty, 'op_out': empty, 'pt_out': empty, 'rs_out': empty,
        }

    # No ISS reaproveitamos os campos físicos do ICMS (Base ICMS, Alíq. ICMS).
    # Por clareza, criamos colunas ALIQ_ISS espelhando ALIQ_ICMS para ler o código.
    df['ALIQ_ISS'] = df['ALIQ_ICMS']

    SKEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'IS_INTRA', 'ALIQ_ISS', 'BASE_FORMULA']

    # ── Regra de Base (BIS) ──────────────────────────────────────────────────
    bf_uniq = df['BASE_FORMULA'].drop_duplicates().sort_values().reset_index(drop=True)
    bf_table = pd.DataFrame({'FORMULA': bf_uniq})
    bf_table.insert(0, 'CODIGO DA REGRA', gen_ids('B', 'IS', len(bf_table)))
    bf_table.insert(1, 'TIPO DE BASE', 'ISS')
    bf_map = {r['FORMULA']: r['CODIGO DA REGRA'] for _, r in bf_table.iterrows()}
    df['BIS'] = df['BASE_FORMULA'].map(bf_map)

    # ── Regra de Alíquota (AIS) ───────────────────────────────────────────────
    aq_table = df[['ALIQ_ISS']].drop_duplicates().sort_values('ALIQ_ISS').reset_index(drop=True)
    aq_table.insert(0, 'CODIGO DA REGRA', gen_ids('A', 'IS', len(aq_table)))
    aq_table['ALIQUOTA_DEC'] = aq_table['ALIQ_ISS'].apply(lambda x: round(x / 100, 4))
    aq_map = {r['ALIQ_ISS']: r['CODIGO DA REGRA'] for _, r in aq_table.iterrows()}
    df['AIS'] = df['ALIQ_ISS'].map(aq_map)

    # ── Mestre de cenários ISS ───────────────────────────────────────────────
    mestre = (
        df[SKEY + ['BIS', 'AIS']]
        .drop_duplicates()
        .sort_values(['TIPO_MOV', 'IS_INTRA', 'ALIQ_ISS'])
        .reset_index(drop=True)
    )
    mestre['NUM'] = range(1, len(mestre) + 1)
    mestre['OBS']  = mestre.apply(lambda r: validar_iss(r['ALIQ_ISS'], r['BASE_FORMULA']), axis=1)
    mestre['ERRO'] = mestre['OBS'].str.len() > 0
    mestre['DESCRICAO'] = mestre.apply(lambda r: (
        f"ISS_{r['ALIQ_ISS']}%"
        f"-{'ENT' if r['TIPO_MOV'].lower()=='entrada' else 'SAI'}"
        f"-{'INTRA' if r['IS_INTRA'] else 'INTER'}"
    ), axis=1)

    df = df.merge(mestre[SKEY + ['NUM']], on=SKEY, how='left')

    # ── Perfil Origem/Destino (LIS) — deduplicado ────────────────────────────
    od_df = (df.groupby(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).size()
               .reset_index(name='_')[['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']]
               .drop_duplicates().sort_values(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).reset_index(drop=True))
    od_df, lis_map = dedup_perfis(od_df, 'NUM',
        ['ESTADO_ORIG', 'ESTADO_DEST'], 'LIS', 'LIS')
    od_out = od_df[['LIS', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates() \
        .sort_values(['LIS','ESTADO_ORIG','ESTADO_DEST']).copy()
    od_out.columns = ['CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO']
    mestre['LIS'] = mestre['NUM'].map(lis_map)

    # ── Perfil Operação (OIS) — deduplicado ──────────────────────────────────
    op_df = (df[['NUM', 'CFOP', 'CFOP_DESC', 'TIPO_MOV']].drop_duplicates()
               .sort_values(['NUM', 'CFOP']).reset_index(drop=True))
    op_df, ois_map = dedup_perfis(op_df, 'NUM',
        ['CFOP', 'TIPO_MOV'], 'OIS', 'OIS')
    op_out = op_df[['OIS', 'TIPO_MOV', 'CFOP', 'CFOP_DESC']].drop_duplicates() \
        .sort_values(['OIS','CFOP']).copy()
    op_out.columns = ['CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP']
    mestre['OIS'] = mestre['NUM'].map(ois_map)

    # ── Perfil Participante — deduplicado ────────────────────────────────────
    pt_df = (df[['NUM', 'TIPO_PART', 'CLI_FORN']].drop_duplicates()
               .sort_values(['NUM', 'TIPO_PART', 'CLI_FORN']).reset_index(drop=True))
    pt_df, cis_map = dedup_perfis_participante(pt_df, 'NUM', 'IS')
    pt_out = pt_df[['CIC', 'TIPO_PART', 'CLI_FORN']].drop_duplicates() \
        .sort_values(['CIC','CLI_FORN']).copy()
    pt_out.columns = ['CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE']
    mestre['CIS'] = mestre.apply(
        lambda r: cis_map.get(
            (r['NUM'], 'FORNECEDOR' if r['TIPO_MOV'].lower()=='entrada' else 'CLIENTE'), ''), axis=1)

    # ── Perfil Serviço (RIS) — deduplicado por (Cód. Serviço, Cod. Produto) ──
    rs_df = (df[['NUM', 'COD_SERVICO', 'PRODUTO']].drop_duplicates()
               .sort_values(['NUM', 'COD_SERVICO', 'PRODUTO']).reset_index(drop=True))
    rs_df, ris_map = dedup_perfis(rs_df, 'NUM',
        ['COD_SERVICO', 'PRODUTO'], 'RIS', 'RIS')
    rs_out = rs_df[['RIS', 'COD_SERVICO', 'PRODUTO']].drop_duplicates() \
        .sort_values(['RIS','COD_SERVICO','PRODUTO']).copy()
    rs_out.columns = ['CODIGO DA REGRA', 'CODIGO DE SERVIÇO', 'CODIGO PRODUTO']
    mestre['RIS'] = mestre['NUM'].map(ris_map)

    mestre['CODIGOS_REGRA'] = mestre.apply(
        lambda r: f"{r['BIS']} | {r['AIS']} | {r['LIS']} | {r['OIS']} | {r['CIS']} | {r['RIS']}", axis=1)

    return {
        'mestre':   mestre,
        'df':       df,
        'bf_table': bf_table,
        'aq_table': aq_table,
        'od_out':   od_out,
        'op_out':   op_out,
        'pt_out':   pt_out,
        'rs_out':   rs_out,
    }


# ============================================================
# CONSTRUÇÃO DAS TABELAS PIS/COFINS
# ============================================================

def build_pis_cof_tables(df):
    """
    Constrói todas as tabelas de perfis e regras para PIS/COFINS.

    CHAVE DO CENÁRIO PIS/COF:
      TIPO_MOV + CST_PIS + ALIQ_PIS + CST_COF + ALIQ_COF
    
    ATENÇÃO: usar colunas AMARELAS da SFT (Base PIS, Aliq. PIS, etc.)
    NÃO usar colunas azuis (Base Pis, Base Cofins) = retenção PCC — a fazer depois

    Retorna dicionário com todos os DataFrames de saída e o mestre.
    """
    PKEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'CST_PIS', 'ALIQ_PIS', 'CST_COF', 'ALIQ_COF']

    # ── Alíquota PIS (APC001, APC002) ─────────────────────────────────────────
    api = df[['ALIQ_PIS']].drop_duplicates().sort_values('ALIQ_PIS').reset_index(drop=True)
    api.insert(0, 'CODIGO DA REGRA', gen_ids('A', 'PC', len(api)))
    api['ALIQUOTA_DEC'] = api['ALIQ_PIS'].apply(lambda x: round(x / 100, 4))
    df['APC_PIS'] = df['ALIQ_PIS'].map({r['ALIQ_PIS']: r['CODIGO DA REGRA'] for _, r in api.iterrows()})

    # ── Alíquota COFINS (APC003, APC004 — sequencial continua após PIS) ───────
    n_pis = len(api)
    aci = df[['ALIQ_COF']].drop_duplicates().sort_values('ALIQ_COF').reset_index(drop=True)
    aci.insert(0, 'CODIGO DA REGRA', gen_ids('A', 'PC', len(aci), start=n_pis + 1))
    aci['ALIQUOTA_DEC'] = aci['ALIQ_COF'].apply(lambda x: round(x / 100, 4))
    df['APC_COF'] = df['ALIQ_COF'].map({r['ALIQ_COF']: r['CODIGO DA REGRA'] for _, r in aci.iterrows()})

    # ── CST PIS (EPC001...EPC006) ─────────────────────────────────────────────
    epis = df[['CST_PIS']].drop_duplicates().sort_values('CST_PIS').reset_index(drop=True)
    epis.insert(0, 'CODIGO DA REGRA', gen_ids('E', 'PC', len(epis)))
    epis['CST'] = epis['CST_PIS'].apply(lambda x: str(x).zfill(2))
    epis['DESC'] = epis['CST_PIS'].map(CST_PC_DESC).fillna('Outros')
    df['EPC'] = df['CST_PIS'].map({r['CST_PIS']: r['CODIGO DA REGRA'] for _, r in epis.iterrows()})

    # ── CST COFINS (EPC007...EPC012 — sequencial continua após PIS) ───────────
    n_epis = len(epis)
    ecof = df[['CST_COF']].drop_duplicates().sort_values('CST_COF').reset_index(drop=True)
    ecof.insert(0, 'CODIGO DA REGRA', gen_ids('E', 'PC', len(ecof), start=n_epis + 1))
    ecof['CST'] = ecof['CST_COF'].apply(lambda x: str(x).zfill(2))
    ecof['DESC'] = ecof['CST_COF'].map(CST_PC_DESC).fillna('Outros')
    df['ECC'] = df['CST_COF'].map({r['CST_COF']: r['CODIGO DA REGRA'] for _, r in ecof.iterrows()})

    # ── Mestre PIS/COF ────────────────────────────────────────────────────────
    mestre_pc = (
        df[PKEY + ['APC_PIS', 'APC_COF', 'EPC', 'ECC']].drop_duplicates()
        .sort_values(['TIPO_MOV', 'CST_PIS', 'ALIQ_PIS'])
        .reset_index(drop=True)
    )
    # PKEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'CST_PIS', 'ALIQ_PIS', 'CST_COF', 'ALIQ_COF']
    mestre_pc.columns = ['FILIAL', 'ESTADO_FILIAL', 'TM', 'CP', 'AP', 'CC', 'AC', 'APP', 'APC', 'EPC', 'ECC']
    mestre_pc['NUM_P'] = range(1, len(mestre_pc) + 1)
    mestre_pc['OBS']   = mestre_pc.apply(lambda r: validar_pis_cofins(r['CP'], r['AP'], r['CC'], r['AC']), axis=1)
    mestre_pc['ERRO']  = mestre_pc['OBS'].str.len() > 0
    mestre_pc['DESC']  = mestre_pc.apply(lambda r: (
        f"PIS_{r['AP']}%/COF_{r['AC']}%-CST{str(r['CP']).zfill(2)}"
        f"-{'ENT' if r['TM'].lower()=='entrada' else 'SAI'}"
    ), axis=1)

    # Renomear para merge
    pkey_renamed = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'CST_PIS', 'ALIQ_PIS', 'CST_COF', 'ALIQ_COF']
    mestre_pc_merge = mestre_pc.copy()
    mestre_pc_merge.columns = pkey_renamed + ['APP','APC','EPC','ECC','NUM_P','OBS','ERRO','DESC']
    df = df.merge(mestre_pc_merge[pkey_renamed + ['NUM_P']], on=pkey_renamed, how='left')

    # ── Perfil Origem/Destino PIS/COF (LPC) — deduplicado ────────────────────
    pod_df = (df[['NUM_P', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates()
                .sort_values(['NUM_P', 'ESTADO_ORIG', 'ESTADO_DEST']).reset_index(drop=True))
    pod_df, lpc_map = dedup_perfis(pod_df, 'NUM_P',
        ['ESTADO_ORIG', 'ESTADO_DEST'], 'LPC', 'LPC')
    pod_out = pod_df[['LPC', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates() \
        .sort_values(['LPC','ESTADO_ORIG','ESTADO_DEST']).copy()
    pod_out.columns = ['CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO']
    mestre_pc['LPC'] = mestre_pc['NUM_P'].map(lpc_map)

    # ── Perfil Operação PIS/COF (OPC) — deduplicado ──────────────────────────
    pop_df = (df[['NUM_P', 'CFOP', 'CFOP_DESC', 'TIPO_MOV']].drop_duplicates()
                .sort_values(['NUM_P', 'CFOP']).reset_index(drop=True))
    pop_df, opc_map = dedup_perfis(pop_df, 'NUM_P',
        ['CFOP', 'TIPO_MOV'], 'OPC', 'OPC')
    pop_out = pop_df[['OPC', 'TIPO_MOV', 'CFOP', 'CFOP_DESC']].drop_duplicates() \
        .sort_values(['OPC','CFOP']).copy()
    pop_out.columns = ['CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP']
    mestre_pc['OPC'] = mestre_pc['NUM_P'].map(opc_map)

    # ── Perfil Participante PIS/COF — deduplicado ────────────────────────────
    ppt_df = (df[['NUM_P', 'TIPO_PART', 'CLI_FORN']].drop_duplicates()
                .sort_values(['NUM_P', 'TIPO_PART', 'CLI_FORN']).reset_index(drop=True))
    ppt_df, cpc_map = dedup_perfis_participante(ppt_df, 'NUM_P', 'PC')
    ppt_out = ppt_df[['CIC', 'TIPO_PART', 'CLI_FORN']].drop_duplicates() \
        .sort_values(['CIC','CLI_FORN']).copy()
    ppt_out.columns = ['CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE']
    mestre_pc['CPC'] = mestre_pc.apply(
        lambda r: cpc_map.get(
            (r['NUM_P'], 'FORNECEDOR' if r['TM'].lower()=='entrada' else 'CLIENTE'), ''), axis=1)

    # ── Perfil Produto PIS/COF (PPC) — deduplicado ──────────────────────────
    ppr_df = (df[['NUM_P', 'NCM', 'PRODUTO', 'ORIGEM_PROD']].drop_duplicates()
                .sort_values(['NUM_P', 'NCM', 'PRODUTO', 'ORIGEM_PROD']).reset_index(drop=True))
    ppr_df, ppc_map = dedup_perfis(ppr_df, 'NUM_P',
        ['NCM', 'PRODUTO', 'ORIGEM_PROD'], 'PPC', 'PPC')
    ppr_df['ORIGEM_DESC'] = ppr_df['ORIGEM_PROD'].map(ORIGEM_DESC).fillna('Não identificada')
    ppr_out = ppr_df[['PPC', 'NCM', 'PRODUTO', 'ORIGEM_PROD', 'ORIGEM_DESC']].drop_duplicates() \
        .sort_values(['PPC','NCM','PRODUTO']).copy()
    ppr_out.columns = ['CODIGO DA REGRA', 'NCM PRODUTO', 'CODIGO PRODUTO', 'ORIGEM', 'DESCRIÇÃO ORIGEM']
    mestre_pc['PPC'] = mestre_pc['NUM_P'].map(ppc_map)

    mestre_pc['CODIGOS_REGRA'] = mestre_pc.apply(
        lambda r: f"{r['APP']}|{r['APC']}|{r['EPC']}|{r['ECC']}|{r['LPC']}|{r['OPC']}|{r['CPC']}|{r['PPC']}", axis=1)

    return {
        'mestre_pc': mestre_pc,
        'df':        df,
        'api':       api,
        'aci':       aci,
        'epis':      epis[['CODIGO DA REGRA', 'CST', 'DESC']],
        'ecof':      ecof[['CODIGO DA REGRA', 'CST', 'DESC']],
        'pod_out':   pod_out,
        'pop_out':   pop_out,
        'ppt_out':   ppt_out,
        'ppr_out':   ppr_out,
    }


# ============================================================
# CONSTRUÇÃO DAS TABELAS DE RETENÇÕES (genérica)
# ============================================================

def build_retencao_tables(df, sigla, nome):
    """
    Constrói tabelas de perfis e regras para UMA retenção (PR, CR, SR, IR, NR, TR).

    Filtro de cenário:
      - Para retenções comuns: linha tem retenção se RET_<sigla>_BASE > 0
      - Para ISS Retido (TR):  linha tem retenção se RET_TR_FLAG == 1

    Chave do cenário:
      TIPO_MOV + ALIQ (origem/destino, CFOP, participante, produto agrupados DENTRO)

    Códigos: B<sigla>, A<sigla>, L<sigla>, O<sigla>, F<sigla>/C<sigla>, P<sigla>
    Ex.: para PIS Retido → BPR, APR, LPR, OPR, FPR/CPR, PPR

    Retorna dict com mestre e tabelas auxiliares (vazio se não há linhas).
    """
    # Encontrar config dessa retenção
    cfg = next((c for c in RETENCOES_CONFIG if c[0] == sigla), None)
    if cfg is None:
        raise ValueError(f"Retenção {sigla} não está em RETENCOES_CONFIG")

    col_flag    = cfg[5]
    val_flag    = cfg[6]
    cfop_allow  = cfg[7]
    col_base_field = f'RET_{sigla}_BASE'
    col_aliq_field = f'RET_{sigla}_ALIQ'
    col_valor_field = f'RET_{sigla}_VALOR'

    # Filtro: ISS Retido usa flag; demais usam Base > 0
    if col_flag is not None:
        flag_field = f'RET_{sigla}_FLAG'
        d = df[df[flag_field] == val_flag].copy()
    else:
        d = df[df[col_base_field] > 0].copy()

    # Filtro adicional de CFOPs permitidos (ex.: ISS Retido só com CFOPs de serviço)
    if cfop_allow is not None:
        d = d[d['CFOP'].isin(cfop_allow)].copy()

    d = d.reset_index(drop=True)

    if len(d) == 0:
        empty = pd.DataFrame()
        return {
            'sigla': sigla, 'nome': nome,
            'mestre': empty, 'df': d,
            'bf_table': empty, 'aq_table': empty,
            'od_out': empty, 'op_out': empty, 'pt_out': empty, 'pr_out': empty,
        }

    # Padroniza nome do campo de alíquota para a chave do cenário
    d['ALIQ_RET'] = d[col_aliq_field]

    RKEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'ALIQ_RET']

    # ── Regra de Base (B<sigla>) — retenção tipicamente sem redução ───────────
    # Aqui sempre criamos uma única regra "VALOR DO SERVIÇO/DOCUMENTO" porque
    # a SFT só registra o valor da base efetiva, sem fórmula de redução.
    bf_table = pd.DataFrame({
        'CODIGO DA REGRA': [f'B{sigla}001'],
        'TIPO DE BASE': [nome.upper()],
        'FORMULA': ['BASE INTEGRAL DA RETENÇÃO'],
    })
    bf_map = {'BASE INTEGRAL DA RETENÇÃO': f'B{sigla}001'}
    d['BX'] = f'B{sigla}001'

    # ── Regra de Alíquota (A<sigla>) ──────────────────────────────────────────
    aq_table = d[['ALIQ_RET']].drop_duplicates().sort_values('ALIQ_RET').reset_index(drop=True)
    aq_table.insert(0, 'CODIGO DA REGRA', gen_ids('A', sigla, len(aq_table)))
    aq_table['ALIQUOTA_DEC'] = aq_table['ALIQ_RET'].apply(lambda x: round(x / 100, 4))
    aq_map = {r['ALIQ_RET']: r['CODIGO DA REGRA'] for _, r in aq_table.iterrows()}
    d['AX'] = d['ALIQ_RET'].map(aq_map)

    # ── Mestre de cenários ────────────────────────────────────────────────────
    mestre = (
        d[RKEY + ['BX', 'AX']]
        .drop_duplicates()
        .sort_values(['TIPO_MOV', 'ALIQ_RET'])
        .reset_index(drop=True)
    )
    mestre['NUM'] = range(1, len(mestre) + 1)
    mestre['DESCRICAO'] = mestre.apply(lambda r: (
        f"{nome}_{r['ALIQ_RET']}%"
        f"-{'ENT' if r['TIPO_MOV'].lower()=='entrada' else 'SAI'}"
    ), axis=1)
    mestre['OBS']  = ''   # retenções não têm validações específicas por ora
    mestre['ERRO'] = False

    d = d.merge(mestre[RKEY + ['NUM']], on=RKEY, how='left')

    # ── Perfil Origem/Destino (L<sigla>) — deduplicado ───────────────────────
    od_df = (d.groupby(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).size()
               .reset_index(name='_')[['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']]
               .drop_duplicates().sort_values(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).reset_index(drop=True))
    od_df, lx_map = dedup_perfis(od_df, 'NUM',
        ['ESTADO_ORIG', 'ESTADO_DEST'], 'LX', f'L{sigla}')
    od_out = od_df[['LX', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates() \
        .sort_values(['LX','ESTADO_ORIG','ESTADO_DEST']).copy()
    od_out.columns = ['CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO']
    mestre['LX'] = mestre['NUM'].map(lx_map)

    # ── Perfil Operação (O<sigla>) — deduplicado ─────────────────────────────
    op_df = (d[['NUM', 'CFOP', 'CFOP_DESC', 'TIPO_MOV']].drop_duplicates()
               .sort_values(['NUM', 'CFOP']).reset_index(drop=True))
    op_df, ox_map = dedup_perfis(op_df, 'NUM',
        ['CFOP', 'TIPO_MOV'], 'OX', f'O{sigla}')
    op_out = op_df[['OX', 'TIPO_MOV', 'CFOP', 'CFOP_DESC']].drop_duplicates() \
        .sort_values(['OX','CFOP']).copy()
    op_out.columns = ['CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP']
    mestre['OX'] = mestre['NUM'].map(ox_map)

    # ── Perfil Participante — deduplicado ────────────────────────────────────
    pt_df = (d[['NUM', 'TIPO_PART', 'CLI_FORN']].drop_duplicates()
               .sort_values(['NUM', 'TIPO_PART', 'CLI_FORN']).reset_index(drop=True))
    pt_df, cx_map = dedup_perfis_participante(pt_df, 'NUM', sigla)
    pt_out = pt_df[['CIC', 'TIPO_PART', 'CLI_FORN']].drop_duplicates() \
        .sort_values(['CIC','CLI_FORN']).copy()
    pt_out.columns = ['CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE']
    mestre['CX'] = mestre.apply(
        lambda r: cx_map.get(
            (r['NUM'], 'FORNECEDOR' if r['TIPO_MOV'].lower()=='entrada' else 'CLIENTE'), ''), axis=1)

    # ── Perfil Produto/Serviço (P<sigla>) — deduplicado ──────────────────────
    pr_df = (d[['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']].drop_duplicates()
               .sort_values(['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']).reset_index(drop=True))
    pr_df, px_map = dedup_perfis(pr_df, 'NUM',
        ['NCM', 'PRODUTO', 'ORIGEM_PROD'], 'PX', f'P{sigla}')
    pr_df['ORIGEM_DESC'] = pr_df['ORIGEM_PROD'].map(ORIGEM_DESC).fillna('Não identificada')
    pr_out = pr_df[['PX', 'NCM', 'PRODUTO', 'ORIGEM_PROD', 'ORIGEM_DESC']].drop_duplicates() \
        .sort_values(['PX','NCM','PRODUTO']).copy()
    pr_out.columns = ['CODIGO DA REGRA', 'NCM PRODUTO', 'CODIGO PRODUTO', 'ORIGEM', 'DESCRIÇÃO ORIGEM']
    mestre['PX'] = mestre['NUM'].map(px_map)

    mestre['CODIGOS_REGRA'] = mestre.apply(
        lambda r: f"{r['BX']} | {r['AX']} | {r['LX']} | {r['OX']} | {r['CX']} | {r['PX']}", axis=1)

    return {
        'sigla': sigla, 'nome': nome,
        'mestre':   mestre,
        'df':       d,
        'bf_table': bf_table,
        'aq_table': aq_table,
        'od_out':   od_out,
        'op_out':   op_out,
        'pt_out':   pt_out,
        'pr_out':   pr_out,
    }


# ============================================================
# CONSTRUÇÃO GENÉRICA — IMPOSTOS PLENOS SEM CST
# (PROTEGE GO, DIFAL clássico, Base Destino)
# ============================================================

def build_imposto_pleno_tables(df, sigla, nome, col_aliq_field, col_base_field, filter_mask):
    """
    Constrói tabelas de perfis e regras para um imposto "pleno" sem CST.
    Usado por: PROTEGE GO (PG), ICMS Complementar DIFAL (CD) e Base Destino (CB).

    Estrutura igual ao ICMS porém SEM Regra de Escrituração (não há CST próprio).
    Reusa BASE_FORMULA do ICMS (origem do registro) quando relevante.

    Args:
      df:               DataFrame completo enriquecido pelo process_sft
      sigla:            2 letras (PG, CD, CB)
      nome:             rótulo legível
      col_aliq_field:   nome do campo no df que tem a alíquota deste tributo
      col_base_field:   nome do campo no df que tem a base deste tributo
                        (string vazia '' = usar BASE_FORMULA do ICMS canonizada)
      filter_mask:      máscara booleana do df p/ filtrar linhas com este imposto

    Chave do cenário: TIPO_MOV + IS_INTRA + <ALIQ_X> + <BASE_FORMULA derivada>

    Códigos: B<sigla>, A<sigla>, L<sigla>, O<sigla>, F<sigla>/C<sigla>, P<sigla>
    """
    d = df[filter_mask].copy().reset_index(drop=True)

    if len(d) == 0:
        empty = pd.DataFrame()
        return {
            'sigla': sigla, 'nome': nome,
            'mestre': empty, 'df': d,
            'bf_table': empty, 'aq_table': empty,
            'od_out': empty, 'op_out': empty, 'pt_out': empty, 'pr_out': empty,
        }

    # ALIQ canônica deste imposto
    d['ALIQ_X'] = d[col_aliq_field]

    # BASE_FORMULA — se o imposto tem coluna própria de base, calculamos
    # uma fórmula derivada simples. Caso contrário, herdamos do ICMS.
    if col_base_field and col_base_field in d.columns:
        # Comparar a base do imposto vs BASE_CALC (do ICMS) para detectar redução
        # Para PROTEGE, a base é geralmente bem menor → fica "REDUÇÃO DE X%"
        # Para DIFAL, a Base. Destin é o valor da operação interestadual
        def _bf(row):
            bx = row[col_base_field]
            bi = row['BASE_ICMS']
            if bx <= 0:
                return 'ISENTO / NÃO TRIBUTADO'
            if bi <= 0:
                return 'SEM REDUÇÃO'
            if abs(bx - bi) < 0.02:
                return 'SEM REDUÇÃO'
            if bx > bi:
                return 'SEM REDUÇÃO'  # anomalia
            pct = round((1 - bx / bi) * 100, 2)
            if pct <= 0:
                return 'SEM REDUÇÃO'
            return f'REDUÇÃO DE {pct}%'
        d['BASE_FORMULA_X'] = d.apply(_bf, axis=1)

        # Canonizar por (ALIQ_X) usando mode
        canon = (
            d.groupby(['ALIQ_X'])['BASE_FORMULA_X']
            .agg(lambda x: x.value_counts().index[0])
            .reset_index().rename(columns={'BASE_FORMULA_X': 'BF_CANON'})
        )
        d = d.merge(canon, on=['ALIQ_X'], how='left')
        d['BASE_FORMULA_X'] = d['BF_CANON'].fillna(d['BASE_FORMULA_X'])
        d.drop(columns='BF_CANON', inplace=True)
    else:
        d['BASE_FORMULA_X'] = d['BASE_FORMULA']

    KEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'IS_INTRA', 'ALIQ_X', 'BASE_FORMULA_X']

    # ── Regra de Base ─────────────────────────────────────────────────────────
    bf_uniq = d['BASE_FORMULA_X'].drop_duplicates().sort_values().reset_index(drop=True)
    bf_table = pd.DataFrame({'FORMULA': bf_uniq})
    bf_table.insert(0, 'CODIGO DA REGRA', gen_ids('B', sigla, len(bf_table)))
    bf_table.insert(1, 'TIPO DE BASE', nome.upper())
    bf_map = {r['FORMULA']: r['CODIGO DA REGRA'] for _, r in bf_table.iterrows()}
    d['BX'] = d['BASE_FORMULA_X'].map(bf_map)

    # ── Regra de Alíquota ─────────────────────────────────────────────────────
    aq_table = d[['ALIQ_X']].drop_duplicates().sort_values('ALIQ_X').reset_index(drop=True)
    aq_table.insert(0, 'CODIGO DA REGRA', gen_ids('A', sigla, len(aq_table)))
    aq_table['ALIQUOTA_DEC'] = aq_table['ALIQ_X'].apply(lambda x: round(x / 100, 4))
    aq_map = {r['ALIQ_X']: r['CODIGO DA REGRA'] for _, r in aq_table.iterrows()}
    d['AX'] = d['ALIQ_X'].map(aq_map)

    # ── Mestre ────────────────────────────────────────────────────────────────
    mestre = (
        d[KEY + ['BX', 'AX']]
        .drop_duplicates()
        .sort_values(['TIPO_MOV', 'IS_INTRA', 'ALIQ_X'])
        .reset_index(drop=True)
    )
    mestre['NUM'] = range(1, len(mestre) + 1)
    mestre['DESCRICAO'] = mestre.apply(lambda r: (
        f"{nome}_{r['ALIQ_X']}%"
        f"-{'ENT' if r['TIPO_MOV'].lower()=='entrada' else 'SAI'}"
        f"-{'INTRA' if r['IS_INTRA'] else 'INTER'}"
    ), axis=1)
    mestre['OBS']  = ''
    mestre['ERRO'] = False

    d = d.merge(mestre[KEY + ['NUM']], on=KEY, how='left')

    # ── Perfil Origem/Destino — deduplicado ──────────────────────────────────
    od_df = (d.groupby(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).size()
               .reset_index(name='_')[['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']]
               .drop_duplicates().sort_values(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).reset_index(drop=True))
    od_df, lx_map = dedup_perfis(od_df, 'NUM',
        ['ESTADO_ORIG', 'ESTADO_DEST'], 'LX', f'L{sigla}')
    od_out = od_df[['LX', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates() \
        .sort_values(['LX','ESTADO_ORIG','ESTADO_DEST']).copy()
    od_out.columns = ['CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO']
    mestre['LX'] = mestre['NUM'].map(lx_map)

    # ── Perfil Operação — deduplicado ────────────────────────────────────────
    op_df = (d[['NUM', 'CFOP', 'CFOP_DESC', 'TIPO_MOV']].drop_duplicates()
               .sort_values(['NUM', 'CFOP']).reset_index(drop=True))
    op_df, ox_map = dedup_perfis(op_df, 'NUM',
        ['CFOP', 'TIPO_MOV'], 'OX', f'O{sigla}')
    op_out = op_df[['OX', 'TIPO_MOV', 'CFOP', 'CFOP_DESC']].drop_duplicates() \
        .sort_values(['OX','CFOP']).copy()
    op_out.columns = ['CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP']
    mestre['OX'] = mestre['NUM'].map(ox_map)

    # ── Perfil Participante — deduplicado ────────────────────────────────────
    pt_df = (d[['NUM', 'TIPO_PART', 'CLI_FORN']].drop_duplicates()
               .sort_values(['NUM', 'TIPO_PART', 'CLI_FORN']).reset_index(drop=True))
    pt_df, cx_map = dedup_perfis_participante(pt_df, 'NUM', sigla)
    pt_out = pt_df[['CIC', 'TIPO_PART', 'CLI_FORN']].drop_duplicates() \
        .sort_values(['CIC','CLI_FORN']).copy()
    pt_out.columns = ['CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE']
    mestre['CX'] = mestre.apply(
        lambda r: cx_map.get(
            (r['NUM'], 'FORNECEDOR' if r['TIPO_MOV'].lower()=='entrada' else 'CLIENTE'), ''), axis=1)

    # ── Perfil Produto — deduplicado ─────────────────────────────────────────
    pr_df = (d[['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']].drop_duplicates()
               .sort_values(['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']).reset_index(drop=True))
    pr_df, px_map = dedup_perfis(pr_df, 'NUM',
        ['NCM', 'PRODUTO', 'ORIGEM_PROD'], 'PX', f'P{sigla}')
    pr_df['ORIGEM_DESC'] = pr_df['ORIGEM_PROD'].map(ORIGEM_DESC).fillna('Não identificada')
    pr_out = pr_df[['PX', 'NCM', 'PRODUTO', 'ORIGEM_PROD', 'ORIGEM_DESC']].drop_duplicates() \
        .sort_values(['PX','NCM','PRODUTO']).copy()
    pr_out.columns = ['CODIGO DA REGRA', 'NCM PRODUTO', 'CODIGO PRODUTO', 'ORIGEM', 'DESCRIÇÃO ORIGEM']
    mestre['PX'] = mestre['NUM'].map(px_map)

    mestre['CODIGOS_REGRA'] = mestre.apply(
        lambda r: f"{r['BX']} | {r['AX']} | {r['LX']} | {r['OX']} | {r['CX']} | {r['PX']}", axis=1)

    return {
        'sigla': sigla, 'nome': nome,
        'mestre':   mestre,
        'df':       d,
        'bf_table': bf_table,
        'aq_table': aq_table,
        'od_out':   od_out,
        'op_out':   op_out,
        'pt_out':   pt_out,
        'pr_out':   pr_out,
    }


# ============================================================
# CONSTRUÇÃO DAS TABELAS IPI (com CST = Trib. IPI)
# ============================================================

# Mapa CST IPI (Tabela 4.5.1 do Sped Fiscal)
CST_IPI_DESC = {
    0:  'Entrada com recuperação de crédito',
    1:  'Entrada tributada com alíquota zero',
    2:  'Entrada isenta',
    3:  'Entrada não tributada',
    4:  'Entrada imune',
    5:  'Entrada com suspensão',
    49: 'Outras entradas',
    50: 'Saída tributada',
    51: 'Saída tributada com alíquota zero',
    52: 'Saída isenta',
    53: 'Saída não tributada',
    54: 'Saída imune',
    55: 'Saída com suspensão',
    99: 'Outras saídas',
}


def build_ipi_tables(df):
    """
    Constrói tabelas de perfis e regras para o IPI.

    Filtro: TEM_IPI == True (Trib. IPI preenchido na SFT)

    Chave do cenário: TIPO_MOV + IS_INTRA + TRIB_IPI + ALIQ_IPI + BASE_FORMULA_IPI
    (sem CFOP — CFOP fica DENTRO do Perfil Operação, igual ICMS)

    Códigos: BPI, API, EPI, LPI, OPI, FPI/CPI, PPI
    """
    d = df[df['TEM_IPI']].copy().reset_index(drop=True)

    if len(d) == 0:
        empty = pd.DataFrame()
        return {
            'mestre': empty, 'df': d,
            'bf_table': empty, 'aq_table': empty, 'esc_table': empty,
            'od_out': empty, 'op_out': empty, 'pt_out': empty, 'pr_out': empty,
        }

    sigla = 'PI'
    nome  = 'IPI'

    # BASE_FORMULA do IPI
    # Como a SFT real não tem Vlr Base IPI preenchido, classificamos pelo CST:
    #   CSTs isentos/não-tributados → ISENTO
    #   Demais → SEM REDUÇÃO (mantém íntegro)
    ISENTOS_IPI = {1, 2, 3, 4, 5, 51, 52, 53, 54, 55}
    def _bf_ipi(row):
        if row['TRIB_IPI'] in ISENTOS_IPI:
            return 'ISENTO / NÃO TRIBUTADO'
        if row['ALIQ_IPI'] == 0 and row['BASE_IPI'] == 0:
            return 'ISENTO / NÃO TRIBUTADO'
        if row['BASE_IPI'] <= 0 and row['OUTR_IPI'] > 0:
            return 'ISENTO / NÃO TRIBUTADO'  # foi classificado em "Outras"
        return 'SEM REDUÇÃO'
    d['BASE_FORMULA_PI'] = d.apply(_bf_ipi, axis=1)

    IKEY = ['FILIAL', 'ESTADO_FILIAL', 'TIPO_MOV', 'IS_INTRA', 'TRIB_IPI', 'ALIQ_IPI', 'BASE_FORMULA_PI']

    # ── Regra de Base ─────────────────────────────────────────────────────────
    bf_uniq = d['BASE_FORMULA_PI'].drop_duplicates().sort_values().reset_index(drop=True)
    bf_table = pd.DataFrame({'FORMULA': bf_uniq})
    bf_table.insert(0, 'CODIGO DA REGRA', gen_ids('B', sigla, len(bf_table)))
    bf_table.insert(1, 'TIPO DE BASE', 'IPI')
    bf_map = {r['FORMULA']: r['CODIGO DA REGRA'] for _, r in bf_table.iterrows()}
    d['BPI'] = d['BASE_FORMULA_PI'].map(bf_map)

    # ── Regra de Alíquota ─────────────────────────────────────────────────────
    aq_table = d[['ALIQ_IPI']].drop_duplicates().sort_values('ALIQ_IPI').reset_index(drop=True)
    aq_table.insert(0, 'CODIGO DA REGRA', gen_ids('A', sigla, len(aq_table)))
    aq_table['ALIQUOTA_DEC'] = aq_table['ALIQ_IPI'].apply(lambda x: round(x / 100, 4))
    aq_map = {r['ALIQ_IPI']: r['CODIGO DA REGRA'] for _, r in aq_table.iterrows()}
    d['API'] = d['ALIQ_IPI'].map(aq_map)

    # ── Regra de Escrituração — CST IPI ──────────────────────────────────────
    esc_table = d[['TRIB_IPI']].drop_duplicates().sort_values('TRIB_IPI').reset_index(drop=True)
    esc_table.insert(0, 'CODIGO DA REGRA', gen_ids('E', sigla, len(esc_table)))
    esc_table['CST IPI'] = esc_table['TRIB_IPI'].apply(lambda x: str(int(x)).zfill(2))
    esc_table['DESCRIÇÃO'] = esc_table['TRIB_IPI'].map(CST_IPI_DESC).fillna('Outros')
    esc_map = {r['TRIB_IPI']: r['CODIGO DA REGRA'] for _, r in esc_table.iterrows()}
    d['EPI'] = d['TRIB_IPI'].map(esc_map)

    # ── Mestre ────────────────────────────────────────────────────────────────
    mestre = (
        d[IKEY + ['BPI', 'API', 'EPI']]
        .drop_duplicates()
        .sort_values(['TIPO_MOV', 'IS_INTRA', 'TRIB_IPI', 'ALIQ_IPI'])
        .reset_index(drop=True)
    )
    mestre['NUM'] = range(1, len(mestre) + 1)
    mestre['DESCRICAO'] = mestre.apply(lambda r: (
        f"IPI_{r['ALIQ_IPI']}%-CST{str(int(r['TRIB_IPI'])).zfill(2)}"
        f"-{'ENT' if r['TIPO_MOV'].lower()=='entrada' else 'SAI'}"
        f"-{'INTRA' if r['IS_INTRA'] else 'INTER'}"
    ), axis=1)
    mestre['OBS']  = ''
    mestre['ERRO'] = False

    d = d.merge(mestre[IKEY + ['NUM']], on=IKEY, how='left')

    # ── Perfil Origem/Destino (LPI) — deduplicado ───────────────────────────
    od_df = (d.groupby(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).size()
               .reset_index(name='_')[['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']]
               .drop_duplicates().sort_values(['NUM', 'ESTADO_ORIG', 'ESTADO_DEST']).reset_index(drop=True))
    od_df, lpi_map = dedup_perfis(od_df, 'NUM',
        ['ESTADO_ORIG', 'ESTADO_DEST'], 'LPI', 'LPI')
    od_out = od_df[['LPI', 'ESTADO_ORIG', 'ESTADO_DEST']].drop_duplicates() \
        .sort_values(['LPI','ESTADO_ORIG','ESTADO_DEST']).copy()
    od_out.columns = ['CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO']
    mestre['LPI'] = mestre['NUM'].map(lpi_map)

    # ── Perfil Operação (OPI) — deduplicado ─────────────────────────────────
    op_df = (d[['NUM', 'CFOP', 'CFOP_DESC', 'TIPO_MOV']].drop_duplicates()
               .sort_values(['NUM', 'CFOP']).reset_index(drop=True))
    op_df, opi_map = dedup_perfis(op_df, 'NUM',
        ['CFOP', 'TIPO_MOV'], 'OPI', 'OPI')
    op_out = op_df[['OPI', 'TIPO_MOV', 'CFOP', 'CFOP_DESC']].drop_duplicates() \
        .sort_values(['OPI','CFOP']).copy()
    op_out.columns = ['CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP']
    mestre['OPI'] = mestre['NUM'].map(opi_map)

    # ── Perfil Participante — deduplicado ───────────────────────────────────
    pt_df = (d[['NUM', 'TIPO_PART', 'CLI_FORN']].drop_duplicates()
               .sort_values(['NUM', 'TIPO_PART', 'CLI_FORN']).reset_index(drop=True))
    pt_df, cpi_map = dedup_perfis_participante(pt_df, 'NUM', 'PI')
    pt_out = pt_df[['CIC', 'TIPO_PART', 'CLI_FORN']].drop_duplicates() \
        .sort_values(['CIC','CLI_FORN']).copy()
    pt_out.columns = ['CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE']
    mestre['CPI'] = mestre.apply(
        lambda r: cpi_map.get(
            (r['NUM'], 'FORNECEDOR' if r['TIPO_MOV'].lower()=='entrada' else 'CLIENTE'), ''), axis=1)

    # ── Perfil Produto (PPI) — deduplicado ──────────────────────────────────
    pr_df = (d[['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']].drop_duplicates()
               .sort_values(['NUM', 'NCM', 'PRODUTO', 'ORIGEM_PROD']).reset_index(drop=True))
    pr_df, ppi_map = dedup_perfis(pr_df, 'NUM',
        ['NCM', 'PRODUTO', 'ORIGEM_PROD'], 'PPI', 'PPI')
    pr_df['ORIGEM_DESC'] = pr_df['ORIGEM_PROD'].map(ORIGEM_DESC).fillna('Não identificada')
    pr_out = pr_df[['PPI', 'NCM', 'PRODUTO', 'ORIGEM_PROD', 'ORIGEM_DESC']].drop_duplicates() \
        .sort_values(['PPI','NCM','PRODUTO']).copy()
    pr_out.columns = ['CODIGO DA REGRA', 'NCM PRODUTO', 'CODIGO PRODUTO', 'ORIGEM', 'DESCRIÇÃO ORIGEM']
    mestre['PPI'] = mestre['NUM'].map(ppi_map)

    mestre['CODIGOS_REGRA'] = mestre.apply(
        lambda r: f"{r['BPI']} | {r['API']} | {r['LPI']} | {r['EPI']} | {r['OPI']} | {r['CPI']} | {r['PPI']}", axis=1)

    return {
        'mestre':   mestre,
        'df':       d,
        'bf_table': bf_table,
        'aq_table': aq_table,
        'esc_table': esc_table[['CODIGO DA REGRA', 'CST IPI', 'DESCRIÇÃO']],
        'od_out':   od_out,
        'op_out':   op_out,
        'pt_out':   pt_out,
        'pr_out':   pr_out,
    }


# ============================================================
# GERAÇÃO DO EXCEL
# ============================================================

def write_excel(icms_tables, pc_tables, iss_tables, ret_tables_list,
                pg_tables, cd_tables, cb_tables, ipi_tables,
                modelo_file, output_file, valor_declaratorio_df=None):
    """
    Escreve a planilha de saída em layout CONSOLIDADO (10 abas).

    Cada aba inclui uma coluna 'TRIBUTO' que identifica de qual imposto
    cada linha pertence, com cor de fundo coerente.
    """

    wb = load_workbook(modelo_file)

    # ── Estilos por tributo ────────────────────────────────────────────────
    # Tupla: (header_color, alt_row_color)
    THEME = {
        'ICMS':              ('1F4E79', 'D6E4F0'),  # Azul
        'PIS':               ('375623', 'E2EFDA'),  # Verde
        'COFINS':            ('375623', 'E2EFDA'),
        'PIS/COFINS':        ('375623', 'E2EFDA'),
        'ISS':               ('C65911', 'FCE4D6'),  # Laranja
        'PIS Retido':        ('5B2C6F', 'E5D5EB'),  # Roxo Retenções
        'COFINS Retido':     ('5B2C6F', 'E5D5EB'),
        'CSLL Retido':       ('5B2C6F', 'E5D5EB'),
        'IRRF':              ('5B2C6F', 'E5D5EB'),
        'INSS':              ('5B2C6F', 'E5D5EB'),
        'ISS Retido':        ('5B2C6F', 'E5D5EB'),
        'PROTEGE GO':        ('C00076', 'FBE4ED'),  # Rosa
        'ICMS Compl. DIFAL': ('117A8B', 'D4EBEF'),  # Ciano
        'ICMS Compl. Base Destino': ('117A8B', 'D4EBEF'),
        'IPI':               ('B7950B', 'F4ECC8'),  # Mostarda
    }
    NEUTRO_HDR = '404040'

    # ── Tipografia / bordas ────────────────────────────────────────────────
    HF  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    DF  = Font(name='Arial', size=9)
    EF  = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    THIN = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    ER  = PatternFill('solid', start_color='C00000')

    def fill_for(tributo):
        color = THEME.get(tributo, (NEUTRO_HDR, 'EEEEEE'))[1]
        return PatternFill('solid', start_color=color)

    def clear_sheet(ws):
        for row in ws.iter_rows():
            for c in row:
                c.value = None; c.fill = PatternFill(); c.font = Font()
                c.border = Border(); c.alignment = Alignment()
        ws.auto_filter.ref = None

    def write_consolidated(ws, headers, rows, col_widths, error_col_idx=None,
                           tributo_col_idx=1):
        """
        Escreve aba consolidada.
        - rows: lista de tuplas, onde o item em tributo_col_idx (1-based) é o tributo
        - error_col_idx: se setado e bool True, linha em vermelho
        """
        clear_sheet(ws)
        ws.row_dimensions[1].height = 32

        # Cabeçalho — cinza escuro neutro (aba unificada)
        hdr_fill = PatternFill('solid', start_color=NEUTRO_HDR)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill; cell.font = HF; cell.alignment = CTR; cell.border = THIN

        # Dados
        for ri, row in enumerate(rows, 2):
            tributo = row[tributo_col_idx - 1] if len(row) >= tributo_col_idx else ''
            row_fill = fill_for(str(tributo))
            has_error = bool(row[error_col_idx - 1]) if error_col_idx and len(row) >= error_col_idx else False
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = THIN
                if has_error:
                    cell.fill = ER; cell.font = EF
                else:
                    cell.fill = row_fill; cell.font = DF
                cell.alignment = CTR if ci <= 2 else LFT

        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row,1)}"

    def get_or_create_sheet(wb, name):
        if name not in wb.sheetnames:
            wb.create_sheet(name)
        return wb[name]

    # ── Coletar registros por tipo de aba ──────────────────────────────────
    perfis_od   = []  # (TRIBUTO, CODIGO, ESTADO_ORIGEM, ESTADO_DESTINO)
    perfis_pt   = []  # (TRIBUTO, CODIGO, TIPO_PART, COD_PART)
    perfis_op   = []  # (TRIBUTO, CODIGO, TIPO_OP, COD_OP, DESC_CFOP)
    perfis_prod = []  # (TRIBUTO, CODIGO, NCM, CODIGO_PRODUTO, ORIGEM, DESC_ORIGEM, COD_SERVICO)
    regras_bf   = []  # (TRIBUTO, CODIGO, TIPO_BASE, FORMULA)
    regras_aq   = []  # (TRIBUTO, CODIGO, ALIQ%, ALIQ_DEC)
    regras_esc  = []  # (TRIBUTO, CODIGO, CST, DESCRIÇÃO)
    mestres     = []  # (TRIBUTO, #, ESTADO_FILIAL, TIPO_MOV, INTRA/INTER, CST/CSTs, BASE_FORMULA,
                      #  ALIQUOTA, PERF_OD, PERF_OP, PERF_PART, PERF_PROD, COD_BASE, COD_ALIQ,
                      #  COD_ESCRIT, DETALHES, CODIGOS_REGRA, DESCRIÇÃO, OBSERVAÇÃO, ⚠)

    # ── Helper para adicionar registros padronizados ───────────────────────
    def _add_od(tributo, df_out):
        for _, r in df_out.iterrows():
            perfis_od.append((tributo, r['CODIGO DA REGRA'], r['ESTADO ORIGEM'], r['ESTADO DESTINO']))

    def _add_pt(tributo, df_out):
        for _, r in df_out.iterrows():
            perfis_pt.append((tributo, r['CODIGO DA REGRA'], r['TIPO PARTICIPANTE'], r['CODIGO PARTICIPANTE']))

    def _add_op(tributo, df_out):
        for _, r in df_out.iterrows():
            perfis_op.append((tributo, r['CODIGO DA REGRA'], r['TIPO DE OPERAÇÃO'],
                              r['CODIGO OPERAÇÃO'], r['DESCRIÇÃO CFOP']))

    def _add_prod(tributo, df_out, kind='prod'):
        """kind: 'prod' (NCM+Prod+Origem) ou 'serv' (Cód.Serviço+Prod)"""
        for _, r in df_out.iterrows():
            if kind == 'serv':
                perfis_prod.append((tributo, r['CODIGO DA REGRA'], '', r['CODIGO PRODUTO'],
                                    '', '', r.get('CODIGO DE SERVIÇO','')))
            else:
                perfis_prod.append((tributo, r['CODIGO DA REGRA'], r.get('NCM PRODUTO',''),
                                    r.get('CODIGO PRODUTO',''), r.get('ORIGEM',''),
                                    r.get('DESCRIÇÃO ORIGEM',''), ''))

    def _add_bf(tributo, df_out):
        for _, r in df_out.iterrows():
            regras_bf.append((tributo, r['CODIGO DA REGRA'], r['TIPO DE BASE'], r['FORMULA']))

    def _add_aq(tributo, df_out, aliq_col='ALIQ_X'):
        for _, r in df_out.iterrows():
            regras_aq.append((tributo, r['CODIGO DA REGRA'],
                              f"{r[aliq_col]}%", r['ALIQUOTA_DEC']))

    def _add_esc(tributo, df_out, cst_col='CST'):
        for _, r in df_out.iterrows():
            cst = r.get(cst_col, r.get('CST ICMS', r.get('CST IPI', '')))
            desc = r.get('DESCRIÇÃO', r.get('DESC', ''))
            regras_esc.append((tributo, r['CODIGO DA REGRA'], cst, desc))

    # ════════════════════════════════════════════════════════════════════════
    # 1) ICMS
    # ════════════════════════════════════════════════════════════════════════
    ic = icms_tables
    mi = ic['mestre']
    _add_od('ICMS', ic['od_out'])
    _add_pt('ICMS', ic['pt_out'])
    _add_op('ICMS', ic['op_out'])
    _add_prod('ICMS', ic['pr_out'])
    _add_bf('ICMS', ic['bf_table'])
    # Alíq: a tabela aq_table do ICMS tem coluna ALIQ_ICMS, não ALIQ_X
    for _, r in ic['aq_table'].iterrows():
        regras_aq.append(('ICMS', r['CODIGO DA REGRA'], f"{r['ALIQ_ICMS']}%", r['ALIQUOTA_DEC']))
    # Esc: tabela tem 'CST ICMS'
    for _, r in ic['esc_table'].iterrows():
        regras_esc.append(('ICMS', r['CODIGO DA REGRA'], r['CST ICMS'], ''))
    # Mestre
    for _, r in mi.iterrows():
        mestres.append((
            'ICMS', r['NUM'], r.get('FILIAL',''), r['ESTADO_FILIAL'], r['TIPO_MOV'],
            'INTRA' if r['IS_INTRA'] else 'INTER',
            str(r['SIT_TRIB']).zfill(2),
            r['BASE_FORMULA'], f"{r['ALIQ_ICMS']}%",
            r['LIC'], r['OIC'], r['CIC'], r['PIC'],
            r['BIC'], r['AIC'], r['EIC'],
            '',  # detalhes
            r['CODIGOS_REGRA'], r['DESCRICAO'], r['OBS'], r['ERRO'],
        ))

    # ════════════════════════════════════════════════════════════════════════
    # 2) ISS
    # ════════════════════════════════════════════════════════════════════════
    mis_iss = iss_tables.get('mestre', pd.DataFrame())
    if len(mis_iss) > 0:
        _add_od('ISS', iss_tables['od_out'])
        _add_pt('ISS', iss_tables['pt_out'])
        _add_op('ISS', iss_tables['op_out'])
        _add_prod('ISS', iss_tables['rs_out'], kind='serv')
        _add_bf('ISS', iss_tables['bf_table'])
        for _, r in iss_tables['aq_table'].iterrows():
            regras_aq.append(('ISS', r['CODIGO DA REGRA'], f"{r['ALIQ_ISS']}%", r['ALIQUOTA_DEC']))
        # ISS sem CST → sem escrituração
        for _, r in mis_iss.iterrows():
            mestres.append((
                'ISS', r['NUM'], r.get('FILIAL',''), r['ESTADO_FILIAL'], r['TIPO_MOV'],
                'INTRA' if r['IS_INTRA'] else 'INTER',
                '',  # sem CST
                r['BASE_FORMULA'], f"{r['ALIQ_ISS']}%",
                r['LIS'], r['OIS'], r['CIS'], r['RIS'],
                r['BIS'], r['AIS'], '',
                '',  # detalhes
                r['CODIGOS_REGRA'], r['DESCRICAO'], r['OBS'], r['ERRO'],
            ))

    # ════════════════════════════════════════════════════════════════════════
    # 3) PIS/COFINS — separados em dois tributos pois cada um tem sua alíq/CST
    # ════════════════════════════════════════════════════════════════════════
    pc = pc_tables
    mpc = pc['mestre_pc']
    # Perfis são compartilhados → marcamos como 'PIS/COFINS'
    _add_od('PIS/COFINS', pc['pod_out'])
    _add_pt('PIS/COFINS', pc['ppt_out'])
    _add_op('PIS/COFINS', pc['pop_out'])
    _add_prod('PIS/COFINS', pc['ppr_out'])
    # Alíq PIS e COFINS separadas
    for _, r in pc['api'].iterrows():
        regras_aq.append(('PIS', r['CODIGO DA REGRA'], f"{r['ALIQ_PIS']}%", r['ALIQUOTA_DEC']))
    for _, r in pc['aci'].iterrows():
        regras_aq.append(('COFINS', r['CODIGO DA REGRA'], f"{r['ALIQ_COF']}%", r['ALIQUOTA_DEC']))
    # CST PIS e CST COFINS separados
    for _, r in pc['epis'].iterrows():
        regras_esc.append(('PIS', r['CODIGO DA REGRA'], r['CST'], r['DESC']))
    for _, r in pc['ecof'].iterrows():
        regras_esc.append(('COFINS', r['CODIGO DA REGRA'], r['CST'], r['DESC']))
    # Mestre PIS/COF — coluna detalhes leva o resumo PIS/COF
    for _, r in mpc.iterrows():
        cst_combo = f"PIS:{str(r['CP']).zfill(2)}/COF:{str(r['CC']).zfill(2)}"
        detalhes = (
            f"Alíq PIS {r['AP']}% (cod {r['APP']}) | "
            f"Alíq COF {r['AC']}% (cod {r['APC']}) | "
            f"CST PIS {r['EPC']} | CST COF {r['ECC']}"
        )
        # Códigos compostos
        cod_aliq = f"{r['APP']} + {r['APC']}"
        cod_esc  = f"{r['EPC']} + {r['ECC']}"
        mestres.append((
            'PIS/COFINS', r['NUM_P'], r.get('FILIAL',''), r.get('ESTADO_FILIAL',''), r['TM'],
            '',  # intra/inter não se aplica ao PIS/COF
            cst_combo,
            'SEM REDUÇÃO',     # PIS/COF não usa fórmula de base aqui
            f"{r['AP']}% / {r['AC']}%",
            r['LPC'], r['OPC'], r['CPC'], r['PPC'],
            '',  # BX (não há regra de base única no PIS/COF)
            cod_aliq, cod_esc,
            detalhes,
            r['CODIGOS_REGRA'], r['DESC'], r['OBS'], r['ERRO'],
        ))

    # ════════════════════════════════════════════════════════════════════════
    # 4) RETENÇÕES (PR, CR, SR, IR, NR, TR)
    # ════════════════════════════════════════════════════════════════════════
    for rt in ret_tables_list:
        nome  = rt['nome']
        mest  = rt['mestre']
        if len(mest) == 0:
            continue
        _add_od(nome, rt['od_out'])
        _add_pt(nome, rt['pt_out'])
        _add_op(nome, rt['op_out'])
        _add_prod(nome, rt['pr_out'])
        _add_bf(nome, rt['bf_table'])
        for _, r in rt['aq_table'].iterrows():
            regras_aq.append((nome, r['CODIGO DA REGRA'], f"{r['ALIQ_RET']}%", r['ALIQUOTA_DEC']))
        # Retenções sem escrituração
        for _, r in mest.iterrows():
            mestres.append((
                nome, r['NUM'], r.get('FILIAL',''), r['ESTADO_FILIAL'], r['TIPO_MOV'],
                '',  # sem intra/inter para retenção
                '',  # sem CST
                'BASE INTEGRAL DA RETENÇÃO', f"{r['ALIQ_RET']}%",
                r['LX'], r['OX'], r['CX'], r['PX'],
                r['BX'], r['AX'], '',
                '',
                r['CODIGOS_REGRA'], r['DESCRICAO'], '', False,
            ))

    # ════════════════════════════════════════════════════════════════════════
    # 5) IMPOSTOS PLENOS: PROTEGE GO, DIFAL, Base Destino
    # ════════════════════════════════════════════════════════════════════════
    for tbl in [pg_tables, cd_tables, cb_tables]:
        nome  = tbl.get('nome', '')
        mest  = tbl.get('mestre', pd.DataFrame())
        if len(mest) == 0:
            continue
        _add_od(nome, tbl['od_out'])
        _add_pt(nome, tbl['pt_out'])
        _add_op(nome, tbl['op_out'])
        _add_prod(nome, tbl['pr_out'])
        _add_bf(nome, tbl['bf_table'])
        _add_aq(nome, tbl['aq_table'], aliq_col='ALIQ_X')
        for _, r in mest.iterrows():
            mestres.append((
                nome, r['NUM'], r.get('FILIAL',''), r['ESTADO_FILIAL'], r['TIPO_MOV'],
                'INTRA' if r['IS_INTRA'] else 'INTER',
                '',  # sem CST
                r['BASE_FORMULA_X'], f"{r['ALIQ_X']}%",
                r['LX'], r['OX'], r['CX'], r['PX'],
                r['BX'], r['AX'], '',
                '',
                r['CODIGOS_REGRA'], r['DESCRICAO'], r['OBS'], r['ERRO'],
            ))

    # ════════════════════════════════════════════════════════════════════════
    # 6) IPI
    # ════════════════════════════════════════════════════════════════════════
    mipi = ipi_tables.get('mestre', pd.DataFrame())
    if len(mipi) > 0:
        _add_od('IPI', ipi_tables['od_out'])
        _add_pt('IPI', ipi_tables['pt_out'])
        _add_op('IPI', ipi_tables['op_out'])
        _add_prod('IPI', ipi_tables['pr_out'])
        _add_bf('IPI', ipi_tables['bf_table'])
        for _, r in ipi_tables['aq_table'].iterrows():
            regras_aq.append(('IPI', r['CODIGO DA REGRA'], f"{r['ALIQ_IPI']}%", r['ALIQUOTA_DEC']))
        for _, r in ipi_tables['esc_table'].iterrows():
            regras_esc.append(('IPI', r['CODIGO DA REGRA'], r['CST IPI'], r['DESCRIÇÃO']))
        for _, r in mipi.iterrows():
            mestres.append((
                'IPI', r['NUM'], r.get('FILIAL',''), r['ESTADO_FILIAL'], r['TIPO_MOV'],
                'INTRA' if r['IS_INTRA'] else 'INTER',
                str(int(r['TRIB_IPI'])).zfill(2),
                r['BASE_FORMULA_PI'], f"{r['ALIQ_IPI']}%",
                r['LPI'], r['OPI'], r['CPI'], r['PPI'],
                r['BPI'], r['API'], r['EPI'],
                '',
                r['CODIGOS_REGRA'], r['DESCRICAO'], r['OBS'], r['ERRO'],
            ))

    # ════════════════════════════════════════════════════════════════════════
    # ESCRITA — Apagar abas antigas do modelo + criar as 9 consolidadas
    # ════════════════════════════════════════════════════════════════════════
    keep = {
        'PERFIL ORIGEM DESTINO', 'PERFIL PARTICIPANTE', 'PERFIL OPERAÇÃO',
        'PERFIL PRODUTO', 'REGRA DE BASE', 'REGRA DE ALIQUOTA',
        'REGRA DE ESCRITURAÇÃO', 'REGRA DE CALCULO',
        'REGRA VALOR DECLARATÓRIO',
        'RESUMO',
    }
    # Remover abas que existirem no modelo e não estejam no keep
    for s in list(wb.sheetnames):
        if s not in keep:
            del wb[s]

    # 1) PERFIL ORIGEM DESTINO
    ws = get_or_create_sheet(wb, 'PERFIL ORIGEM DESTINO')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'ESTADO ORIGEM', 'ESTADO DESTINO'],
        perfis_od,
        [22, 18, 16, 16])

    # 2) PERFIL PARTICIPANTE
    ws = get_or_create_sheet(wb, 'PERFIL PARTICIPANTE')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'TIPO PARTICIPANTE', 'CODIGO PARTICIPANTE'],
        perfis_pt,
        [22, 18, 20, 24])

    # 3) PERFIL OPERAÇÃO
    ws = get_or_create_sheet(wb, 'PERFIL OPERAÇÃO')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'TIPO DE OPERAÇÃO', 'CODIGO OPERAÇÃO', 'DESCRIÇÃO CFOP'],
        perfis_op,
        [22, 18, 18, 16, 68])

    # 4) PERFIL PRODUTO
    ws = get_or_create_sheet(wb, 'PERFIL PRODUTO')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'NCM PRODUTO', 'CODIGO PRODUTO', 'ORIGEM', 'DESCRIÇÃO ORIGEM', 'CODIGO DE SERVIÇO'],
        perfis_prod,
        [22, 18, 14, 22, 8, 42, 22])

    # 5) REGRA DE BASE
    ws = get_or_create_sheet(wb, 'REGRA DE BASE')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'TIPO DE BASE', 'FORMULA'],
        regras_bf,
        [22, 18, 22, 34])

    # 6) REGRA DE ALIQUOTA
    ws = get_or_create_sheet(wb, 'REGRA DE ALIQUOTA')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'ALÍQUOTA (%)', 'ALÍQUOTA DECIMAL'],
        regras_aq,
        [22, 18, 16, 18])

    # 7) REGRA DE ESCRITURAÇÃO
    ws = get_or_create_sheet(wb, 'REGRA DE ESCRITURAÇÃO')
    write_consolidated(ws,
        ['TRIBUTO', 'CODIGO DA REGRA', 'CST (2 dígitos)', 'DESCRIÇÃO'],
        regras_esc,
        [22, 18, 16, 52])

    # 8) REGRA DE CALCULO (mestre consolidado)
    ws = get_or_create_sheet(wb, 'REGRA DE CALCULO')
    write_consolidated(ws,
        ['TRIBUTO', '#', 'FILIAL', 'ESTADO FILIAL', 'TIPO MOV.', 'INTRA/INTER', 'CST',
         'BASE CÁLCULO', 'ALÍQUOTA',
         'Perf. Ori/Des', 'Perf. Oper.', 'Perf. Part.', 'Perf. Prod./Serv.',
         'Cod. Base', 'Cod. Alíq', 'Cod. Escrit.',
         'DETALHES', 'CÓDIGOS DA REGRA', 'DESCRIÇÃO',
         'OBSERVAÇÃO / INCONSISTÊNCIA', '⚠'],
        mestres,
        [22, 5, 8, 13, 12, 10, 14, 28, 16, 14, 14, 14, 14, 12, 12, 12, 50, 50, 38, 50, 5],
        error_col_idx=21)

    # 9) RESUMO
    resumo_rows = []
    resumo_rows.append(('ICMS', len(mi), mi['ERRO'].sum(),
                        f"Chave: TIPO_MOV+INTRA+CST+ALIQ+BASE_FORMULA. CFOPs de serviço (1933/2933/5933/6933) desviados para ISS."))
    if len(mis_iss) > 0:
        resumo_rows.append(('ISS', len(mis_iss), mis_iss['ERRO'].sum(),
                            "Chave: TIPO_MOV+INTRA+ALIQ+BASE_FORMULA. Sem CST. CFOPs 1933/2933/5933/6933."))
    resumo_rows.append(('PIS/COFINS', len(mpc), mpc['ERRO'].sum(),
                        "Chave: TIPO_MOV+CST_PIS+ALIQ_PIS+CST_COF+ALIQ_COF. Tributação."))
    for rt in ret_tables_list:
        if len(rt['mestre']) > 0:
            resumo_rows.append((rt['nome'], len(rt['mestre']), int(rt['mestre']['ERRO'].sum()),
                                f"Retenção. {len(rt['df'])} linhas SFT com retenção efetiva."))
    if len(pg_tables['mestre']) > 0:
        resumo_rows.append(('PROTEGE GO', len(pg_tables['mestre']), 0,
                            f"Aliq.PROT.GO > 0. {len(pg_tables['df'])} linhas."))
    if len(cd_tables['mestre']) > 0:
        resumo_rows.append(('ICMS Compl. DIFAL', len(cd_tables['mestre']), 0,
                            f"Difal ICMS > 0. {len(cd_tables['df'])} linhas."))
    if len(cb_tables['mestre']) > 0:
        resumo_rows.append(('ICMS Compl. Base Destino', len(cb_tables['mestre']), 0,
                            f"Base. Destin > 0. {len(cb_tables['df'])} linhas."))
    if len(mipi) > 0:
        resumo_rows.append(('IPI', len(mipi), 0,
                            f"Trib. IPI usado como CST. {len(ipi_tables['df'])} linhas."))

    ws = get_or_create_sheet(wb, 'RESUMO')
    write_consolidated(ws,
        ['TRIBUTO', 'CENÁRIOS', 'INCONSISTÊNCIAS', 'OBSERVAÇÕES'],
        resumo_rows,
        [28, 12, 16, 90])

    # 10) REGRA VALOR DECLARATÓRIO (CBENEF cruzado com SFT + NCM + Aplicabilidade)
    ws = get_or_create_sheet(wb, 'REGRA VALOR DECLARATÓRIO')
    vd_rows = []
    if valor_declaratorio_df is not None and len(valor_declaratorio_df) > 0:
        for _, r in valor_declaratorio_df.iterrows():
            vd_rows.append((
                r['TRIBUTO'],
                r['FILIAL'],
                r['CÓDIGO VALOR DECLARATÓRIO'],
                r['CST'],
                r['CFOP'],
                r['TIPO BENEFÍCIO'],
                r['DESCRIÇÃO CBENEF'],
                r.get('NCMs DA SFT', ''),
                r.get('APLICÁVEL?', ''),
                r.get('OBSERVAÇÃO', ''),
            ))
    write_consolidated(ws,
        ['TRIBUTO', 'FILIAL', 'CÓDIGO VALOR DECLARATÓRIO', 'CST', 'CFOP',
         'TIPO BENEFÍCIO', 'DESCRIÇÃO CBENEF', 'NCMs DA SFT',
         'APLICÁVEL?', 'OBSERVAÇÃO'],
        vd_rows,
        [10, 8, 22, 6, 8, 22, 60, 60, 13, 65])

    # Ajustar altura das linhas conforme nº de NCMs
    for row_idx in range(2, ws.max_row + 1):
        ncm_cell = ws.cell(row=row_idx, column=8)
        if ncm_cell.value:
            lines = str(ncm_cell.value).count('\n') + 1
            ws.row_dimensions[row_idx].height = max(15 * lines + 5, 30)

    # Pintar coluna APLICÁVEL? conforme valor: verde SIM, amarelo TALVEZ, cinza NÃO
    fill_sim    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_talvez = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_nao    = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    for row_idx in range(2, ws.max_row + 1):
        cell_apl = ws.cell(row=row_idx, column=9)
        v = str(cell_apl.value or '').strip()
        if v == 'SIM':
            cell_apl.fill = fill_sim
            cell_apl.font = Font(bold=True, color='006100')
        elif v == 'TALVEZ':
            cell_apl.fill = fill_talvez
            cell_apl.font = Font(bold=True, color='9C5700')
        elif v == 'NÃO':
            cell_apl.fill = fill_nao
            cell_apl.font = Font(color='808080')
            # Cinza claro também na linha inteira (visualmente "descartado")
            for col_idx in range(1, 11):
                c = ws.cell(row=row_idx, column=col_idx)
                if not c.fill or c.fill.fgColor.rgb in (None, '00000000'):
                    c.font = Font(color='808080')

    # Ordenar abas
    desired_order = [
        'RESUMO',
        'REGRA DE CALCULO',
        'REGRA VALOR DECLARATÓRIO',
        'PERFIL ORIGEM DESTINO', 'PERFIL PARTICIPANTE',
        'PERFIL OPERAÇÃO', 'PERFIL PRODUTO',
        'REGRA DE BASE', 'REGRA DE ALIQUOTA', 'REGRA DE ESCRITURAÇÃO',
    ]
    sheets_by_name = {s.title: s for s in wb._sheets}
    new_order = [sheets_by_name[n] for n in desired_order if n in sheets_by_name]
    for s in wb._sheets:
        if s not in new_order:
            new_order.append(s)
    wb._sheets = new_order

    wb.save(output_file)
    print(f"Planilha gerada: {output_file}")


def gerar_planilha(sft_path, perfil_empresa=None, sft_sheet=None,
                   log_fn=None, output_path=None):
    """
    Função principal para uso no app web ou linha de comando.

    Args:
      sft_path:        caminho do arquivo SFT (ou objeto file-like)
      perfil_empresa:  dict {flag: bool} do questionário (None = mantém tudo)
      sft_sheet:       nome da aba (None = auto-detectar)
      log_fn:          callable(str) para receber logs. Default: print
      output_path:     caminho onde salvar a planilha. None = retorna bytes

    Retorna:
      Se output_path=None: bytes do arquivo .xlsx gerado
      Se output_path=str:  None (e salva no path)
    """
    import io

    log = log_fn or print
    perfil_empresa = perfil_empresa or {}

    log("=== CONFIGURADOR DE TRIBUTOS PROTHEUS ===")

    log("Carregando CFOP map...")
    cfop_map = load_cfop_map(CFOP_FILE)

    log("Carregando SFT...")
    sft = load_sft(sft_path, sft_sheet)
    log(f"  {len(sft)} linhas carregadas.")

    log("Processando SFT...")
    df = process_sft(sft, cfop_map)
    n_icms = (df['IMPOSTO'] == 'ICMS').sum()
    n_iss  = (df['IMPOSTO'] == 'ISS').sum()
    log(f"  {n_icms} linhas ICMS | {n_iss} linhas ISS (CFOPs {sorted(CFOP_ISS)})")

    log("Carregando tabelas CBENEF...")
    cbenef_tables = load_cbenef_tables(CBENEF_TABLES)
    if not cbenef_tables:
        log("  Nenhuma tabela CBENEF carregada.")

    log("Carregando tabela NCM (descrições)...")
    ncm_map = load_ncm_table(NCM_TABLE_FILE)

    log("Construindo tabelas ICMS...")
    icms_tables = build_icms_tables(df)
    mi = icms_tables['mestre']
    log(f"  {len(mi)} cenários ICMS | {mi['ERRO'].sum()} com inconsistência")

    log("Construindo tabelas ISS...")
    iss_tables = build_iss_tables(df)
    mis = iss_tables['mestre']
    if len(mis) > 0:
        log(f"  {len(mis)} cenários ISS | {mis['ERRO'].sum()} com inconsistência")
    else:
        log("  Nenhuma linha ISS — abas ISS omitidas")

    log("Construindo tabelas PIS/COFINS...")
    pc_tables = build_pis_cof_tables(df)
    mp = pc_tables['mestre_pc']
    log(f"  {len(mp)} cenários PIS/COF | {mp['ERRO'].sum()} com inconsistência")

    log("Construindo tabelas de Retenções...")
    ret_tables_list = []
    for cfg in RETENCOES_CONFIG:
        sigla, nome = cfg[0], cfg[1]
        rt = build_retencao_tables(df, sigla, nome)
        ret_tables_list.append(rt)
        n_cen = len(rt['mestre'])
        n_lin = len(rt['df'])
        marker = '⚠ vazio' if n_cen == 0 else f"{n_cen} cenários, {n_lin} linhas"
        log(f"  {sigla} ({nome:14s}): {marker}")

    log("Construindo tabelas PROTEGE GO (PG)...")
    pg_tables = build_imposto_pleno_tables(
        df, 'PG', 'PROTEGE GO',
        col_aliq_field='ALIQ_PG', col_base_field='BASE_PG',
        filter_mask=df['ALIQ_PG'] > 0,
    )
    log(f"  {len(pg_tables['mestre'])} cenários, {len(pg_tables['df'])} linhas")

    log("Construindo tabelas ICMS Complementar DIFAL (CD)...")
    cd_tables = build_imposto_pleno_tables(
        df, 'CD', 'ICMS Compl. DIFAL',
        col_aliq_field='ALIQ_ICMS', col_base_field='',
        filter_mask=df['DIFAL_ICMS'] > 0,
    )
    log(f"  {len(cd_tables['mestre'])} cenários, {len(cd_tables['df'])} linhas")

    log("Construindo tabelas ICMS Complementar Base Destino (CB)...")
    cb_tables = build_imposto_pleno_tables(
        df, 'CB', 'ICMS Compl. Base Destino',
        col_aliq_field='ALIQ_ICMS', col_base_field='BASE_DESTIN',
        filter_mask=df['BASE_DESTIN'] > 0,
    )
    log(f"  {len(cb_tables['mestre'])} cenários, {len(cb_tables['df'])} linhas")

    log("Construindo tabelas IPI...")
    ipi_tables = build_ipi_tables(df)
    n_cen_ipi = len(ipi_tables['mestre'])
    n_lin_ipi = len(ipi_tables['df'])
    if n_cen_ipi > 0:
        log(f"  {n_cen_ipi} cenários IPI, {n_lin_ipi} linhas")
    else:
        log("  Nenhuma linha IPI — abas IPI omitidas")

    log("Construindo aba REGRA VALOR DECLARATÓRIO (CBENEF × SFT)...")
    valor_declaratorio_df = build_valor_declaratorio_table(
        df, cbenef_tables, ncm_map=ncm_map,
        classificacao_por_estado=CBENEF_CLASSIFICACAO_POR_ESTADO,
        perfil_empresa=perfil_empresa)
    if cbenef_tables:
        n_codigos = len(valor_declaratorio_df)
        n_pares = valor_declaratorio_df[['CST','CFOP']].drop_duplicates().shape[0] if n_codigos else 0
        log(f"  {n_codigos} linhas | {n_pares} pares CST×CFOP únicos")
    else:
        log("  Sem tabelas CBENEF — aba vazia")

    log("Gerando planilha Excel...")
    if output_path:
        # Salvar em disco
        write_excel(icms_tables, pc_tables, iss_tables, ret_tables_list,
                    pg_tables, cd_tables, cb_tables, ipi_tables,
                    MODELO_FILE, output_path,
                    valor_declaratorio_df=valor_declaratorio_df)
        log(f"=== CONCLUÍDO === Arquivo: {output_path}")
        return None
    else:
        # Salvar em buffer e retornar bytes (para Streamlit)
        # write_excel exige path → usamos arquivo temporário
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            write_excel(icms_tables, pc_tables, iss_tables, ret_tables_list,
                        pg_tables, cd_tables, cb_tables, ipi_tables,
                        MODELO_FILE, tmp_path,
                        valor_declaratorio_df=valor_declaratorio_df)
            with open(tmp_path, 'rb') as f:
                data = f.read()
            log(f"=== CONCLUÍDO === {len(data):,} bytes gerados")
            return data
        finally:
            try: os.unlink(tmp_path)
            except OSError: pass


# Permite rodar o script como CLI também
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Configurador de Tributos Protheus')
    parser.add_argument('sft', help='Caminho da SFT.xlsx')
    parser.add_argument('-o', '--output', default='Planilha_Apoio.xlsx',
                        help='Caminho da planilha de saída')
    parser.add_argument('--sheet', default=None, help='Nome da aba SFT (auto se omitido)')
    args = parser.parse_args()

    # Coletar perfil interativamente para CLI
    perfil = coletar_perfil_empresa(PERFIL_PERGUNTAS) if PERFIL_PERGUNTAS else {}
    gerar_planilha(args.sft, perfil_empresa=perfil, sft_sheet=args.sheet,
                   output_path=args.output)
